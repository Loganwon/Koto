# Copyright (C) 2024-2026 Koto AI. All rights reserved.
# SPDX-License-Identifier: LicenseRef-Koto-Proprietary
from __future__ import annotations

import asyncio
import logging
import os
import re
from datetime import datetime

try:
    from google.genai import types
except Exception:  # pragma: no cover - optional SDK in some test envs
    types = None

from web.task_orchestrator_runtime import (
    WORKSPACE_DIR,
    SmartDispatcher,
    client,
    get_filegen_brief_instruction,
    settings_manager,
)
from web.web_searcher import WebSearcher

_app_logger = logging.getLogger("koto.app")


async def execute_file_gen(
    user_input: str,
    context: dict,
    subtask: dict,
    progress_callback=None,
    *,
    ppt_multi_step_runner=None,
) -> dict:
    """执行文件生成子任务
    增强：复杂/长文/要求“深度、详细、研究”时，先运行深度研究并切换到更强模型生成。
    """

    def _report(msg: str, detail: str = ""):
        _app_logger.debug(f"[FILE_GEN] {msg} | {detail}")
        if progress_callback:
            progress_callback(msg, detail)

    try:
        # 提取前一个任务的结果作为输入
        previous_data = context.get(f"step_{subtask['index']}_output", "")

        # 复杂度判定（长文本或显式“深度/详细/研究/全面/技术”请求）
        text_lower = user_input.lower()
        complex_flags = [
            len(user_input) > 120,
            any(
                k in text_lower
                for k in [
                    "深度",
                    "详细",
                    "研究",
                    "全面",
                    "技术",
                    "报告",
                    "综述",
                    "whitepaper",
                ]
            ),
        ]
        is_complex = any(complex_flags)

        # -- Planning Layer (DocumentPlanner) --------------------------
        _doc_plan = None
        if is_complex:
            try:
                from web.doc_planner import DocumentPlanner

                _planner = DocumentPlanner(
                    ai_client=client, model_name="gemini-2.5-flash"
                )
                _report("📋 规划文档结构...", "分析需求/分配章节")
                _doc_plan = await _planner.plan(
                    user_input, previous_context=previous_data
                )
                if _doc_plan.success:
                    _report(
                        f"✅ 规划完成：{len(_doc_plan.sections)} 节 | {_doc_plan.doc_type.upper()}",
                        _doc_plan.to_context_str()[:120],
                    )
                else:
                    _report(
                        "⚠️ 规划层失败，使用默认流程",
                        _doc_plan.error[:60] if _doc_plan.error else "",
                    )
                    _doc_plan = None
            except Exception as _pe:
                _app_logger.warning(f"[FILE_GEN] ⚠️ 规划层异常: {_pe}")
                _doc_plan = None

        # 检测目标格式（PPT、Excel、Word等）
        ppt_keywords = ["ppt", "幻灯片", "演示", "汇报", "presentation", "slide"]
        prefer_ppt = any(kw in user_input.lower() for kw in ppt_keywords)

        prefer_excel = (
            "excel" in user_input.lower()
            or "xlsx" in user_input.lower()
            or "表格" in user_input
        )
        prefer_pdf = "pdf" in user_input.lower()
        if _doc_plan:
            prefer_ppt = (_doc_plan.doc_type == "ppt") or prefer_ppt
            prefer_excel = (_doc_plan.doc_type == "excel") or prefer_excel
            prefer_pdf = (_doc_plan.doc_type == "pdf") or prefer_pdf

        # 根据目标格式选择提示
        if prefer_ppt:
            # 尝试使用新的多阶段生成流程 (Plan-then-Execute)
            try:
                if ppt_multi_step_runner is None:
                    raise RuntimeError("PPT multi-step runner is unavailable")
                ppt_result = await ppt_multi_step_runner(
                    user_input, context, subtask, progress_callback
                )
                if ppt_result.get("success"):
                    _report(
                        f"PPT生成成功",
                        f"文件: {(ppt_result.get('saved_files') or [''])[0]}",
                    )
                    return ppt_result
                elif ppt_result.get("fallback_to_single_step"):
                    _app_logger.warning(
                        "[FILE_GEN] ⚠️ 多阶段生成遇到问题，回退到单步生成逻辑"
                    )
                else:
                    return ppt_result
            except Exception as e:
                _app_logger.warning(f"[FILE_GEN] ⚠️ 多阶段生成异常: {e}")

            # 回退单步提示生成
            gen_prompt = (
                "你是一个顶尖的演示文稿内容策划师和排版规划师。\n\n"
                "在每个 `## 章节标题` 前一行写类型标签来选择幻灯片类型：\n"
                "- `[详细]` — 深入展示 3-5 个要点\n"
                "- `[概览]` — 多主题速览，用 `### 子标题` 分组\n"
                "- `[亮点]` — 关键数据，格式: `- 数值 | 说明`\n"
                "- `[对比]` — 两方对比，用 `### 选项A` 和 `### 选项B` 分组\n"
                "- `[过渡页]` — 章节过渡（最多 2 个）\n\n"
                "**输出格式（严格遵循 Markdown）**：\n"
                "```\n"
                "# 演示标题\n\n"
                "[详细]\n"
                "## 章节标题\n"
                "- 要点1（包含具体信息）\n"
                "- 要点2\n"
                "```\n\n"
                "规则：重点内容用多个 [详细] 展开，简要内容合并到 [概览]，关键数据用 [亮点]。\n"
                "每个要点包含具体信息，中文输出，只输出大纲。\n"
            )
        else:
            if _doc_plan and is_complex:
                # 使用规划层生成增强 prompt（含章节指引）
                from web.doc_planner import build_generation_prompt_from_plan

                gen_prompt = build_generation_prompt_from_plan(
                    _doc_plan, user_input, previous_data
                )
            else:
                gen_prompt = (
                    "你是Koto，一个专业的数据整理与报告生成助手。\n"
                    "请基于用户需求和提供的数据，输出清晰、可直接放入文档的 Markdown 内容。\n"
                    "如果是价格类信息，必须包含一个 Markdown 表格，字段建议为：时间、价格、变化、来源。\n"
                    "输出要求：\n"
                    "- 只输出内容，不要输出代码或 BEGIN_FILE 标记\n"
                    "- 中文输出，结构清晰\n"
                )

        full_input = (
            f"用户原始需求: {context['original_input']}\n\n"
            f"前面步骤的数据/信息:\n{previous_data}\n\n"
            f"{gen_prompt}"
        )

        # 深度研究：为复杂任务先补充研究上下文
        research_context = ""
        if is_complex:
            try:
                research_context = WebSearcher.deep_research_for_ppt(
                    user_input, previous_data
                )
                if research_context:
                    previous_data = f"[深度研究]\n{research_context}\n\n[已有信息]\n{previous_data}"
                    _app_logger.debug(
                        f"[FILE_GEN] 🔬 深度研究完成，追加 {len(research_context)} 字上下文"
                    )
            except Exception as research_err:
                _app_logger.warning(f"[FILE_GEN] ⚠️ 深度研究失败: {research_err}")

        # 调用模型生成内容
        model_id = SmartDispatcher.get_model_for_task(
            "FILE_GEN", complexity="complex" if is_complex else "normal"
        )

        _report(f"正在撰写内容...", f"模型: {model_id}")

        def _generate_text(prompt_text: str) -> str:
            response = client.models.generate_content(
                model=model_id,
                contents=prompt_text,
                config=types.GenerateContentConfig(
                    system_instruction=get_filegen_brief_instruction(),
                    temperature=0.4,
                    max_output_tokens=4000,
                ),
            )
            return response.text or ""

        def _clean_filegen_text(text: str) -> str:
            if not text:
                return text
            cleaned = text

            # Remove fenced code blocks but keep content
            cleaned = re.sub(r"```[a-zA-Z0-9_-]*\n", "", cleaned)
            cleaned = cleaned.replace("```", "")

            # Strip markdown links to plain text
            cleaned = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", cleaned)

            # Remove bold/italic markers
            cleaned = re.sub(r"\*\*(.+?)\*\*", r"\1", cleaned)
            cleaned = re.sub(r"__(.+?)__", r"\1", cleaned)
            cleaned = re.sub(r"\*(.+?)\*", r"\1", cleaned)
            cleaned = re.sub(r"_(.+?)_", r"\1", cleaned)

            # Remove inline code ticks
            cleaned = cleaned.replace("`", "")

            # Strip heading markers and blockquotes at line start
            cleaned = re.sub(r"^\s{0,3}#{1,6}\s+", "", cleaned, flags=re.MULTILINE)
            cleaned = re.sub(r"^\s*>\s?", "", cleaned, flags=re.MULTILINE)

            # Remove horizontal rules
            cleaned = re.sub(r"^\s*[-_*]{3,}\s*$", "", cleaned, flags=re.MULTILINE)

            # Flatten list markers but keep structure via indentation
            cleaned = re.sub(r"^\s*[-*+]\s+", "  ", cleaned, flags=re.MULTILINE)
            cleaned = re.sub(r"^\s*\d+\.\s+", "  ", cleaned, flags=re.MULTILINE)

            # Normalize extra blank lines
            cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)

            # Cleanup leftover marker pairs
            cleaned = cleaned.replace("**", "").replace("__", "")

            return cleaned

        text_out = _generate_text(full_input) or "(无输出)"
        text_out = _clean_filegen_text(text_out)
        _report(f"内容撰写完成", f"共 {len(text_out)} 字")

        # 解析 Markdown 表格
        def _extract_markdown_table(md_text: str):
            lines = [line.strip() for line in md_text.splitlines() if "|" in line]
            for i in range(len(lines) - 1):
                header_line = lines[i]
                sep_line = lines[i + 1]
                if re.match(r"^\s*\|?\s*[-:|\s]+\|\s*$", sep_line):
                    headers = [c.strip() for c in header_line.strip("|").split("|")]
                    rows = []
                    j = i + 2
                    while j < len(lines) and "|" in lines[j]:
                        row = [c.strip() for c in lines[j].strip("|").split("|")]
                        if len(row) < len(headers):
                            row += [""] * (len(headers) - len(row))
                        rows.append(row[: len(headers)])
                        j += 1
                    return [headers] + rows
            return None

        # 解析PPT大纲结构（支持智能规划标签）
        def _parse_ppt_outline(md_text: str) -> dict:
            """解析带 [类型] 标签的 PPT 大纲"""
            lines = md_text.split("\n")
            outline = {"title": "", "slides": []}
            _tmap = {
                "过渡页": "divider",
                "过渡": "divider",
                "详细": "detail",
                "重点": "detail",
                "亮点": "highlight",
                "数据": "highlight",
                "概览": "overview",
                "速览": "overview",
                "简要": "overview",
                "对比": "comparison",
                "比较": "comparison",
            }
            cur_type = "detail"
            cur_slide = None
            cur_sub = None

            for line in lines:
                line = line.rstrip()
                if line.strip() in ("```", "```markdown"):
                    continue
                tm = re.match(r"^\s*\[(.+?)\]\s*$", line)
                if tm:
                    cur_type = _tmap.get(tm.group(1).strip(), "detail")
                    continue
                if line.startswith("# ") and not line.startswith("## "):
                    outline["title"] = line[2:].strip()
                elif line.startswith("## "):
                    if (
                        cur_sub
                        and cur_slide
                        and cur_slide.get("type") in ("overview", "comparison")
                    ):
                        cur_slide.setdefault("subsections", []).append(cur_sub)
                        cur_sub = None
                    if cur_slide:
                        outline["slides"].append(cur_slide)
                    cur_slide = {
                        "type": cur_type,
                        "title": line[3:].strip(),
                        "points": [],
                        "content": [],
                    }
                    if cur_type == "divider":
                        cur_slide["description"] = ""
                    cur_type = "detail"
                    cur_sub = None
                elif line.startswith("### ") and cur_slide:
                    if cur_sub:
                        cur_slide.setdefault("subsections", []).append(cur_sub)
                    cur_sub = {
                        "subtitle": line[4:].strip(),
                        "label": line[4:].strip(),
                        "points": [],
                    }
                elif re.match(r"^[\s]*[-•*]\s", line) and cur_slide is not None:
                    pt = re.sub(r"^[\s]*[-•*]\s+", "", line).strip()
                    if cur_sub is not None:
                        cur_sub["points"].append(pt)
                    else:
                        cur_slide["points"].append(pt)
                        cur_slide["content"].append(pt)
                elif (
                    cur_slide
                    and cur_slide.get("type") == "divider"
                    and line.strip()
                ):
                    cur_slide["description"] = line.strip()

            if (
                cur_sub
                and cur_slide
                and cur_slide.get("type") in ("overview", "comparison")
            ):
                cur_slide.setdefault("subsections", []).append(cur_sub)
            if cur_slide:
                outline["slides"].append(cur_slide)
            for sl in outline["slides"]:
                if sl.get("type") == "comparison" and "subsections" in sl:
                    subs = sl["subsections"]
                    if len(subs) >= 2:
                        sl["left"] = subs[0]
                        sl["right"] = subs[1]
            return outline

        title = "生成文档"
        if "价格" in user_input or "表格" in user_input:
            title = "价格波动表格"
        elif prefer_ppt:
            title = "演示文稿"

        saved_files = []
        file_type = None
        excel_error = None

        # 生成PPT
        if prefer_ppt:
            try:
                from web.ppt_generator import PPTGenerator

                ppt_outline = _parse_ppt_outline(text_out)

                # ── 质量门控 ──
                try:
                    from web.file_quality_checker import FileQualityGate

                    _qg = FileQualityGate.check_and_fix_ppt_outline(
                        ppt_outline.get("slides", []),
                        user_request=user_input,
                        progress_callback=_report,
                    )
                    ppt_outline["slides"] = _qg["outline"]
                except Exception as _qge:
                    _app_logger.warning(f"[FILE_GEN] ⚠️ PPT 质量门控异常: {_qge}")

                # 确定主题（通过关键词检测）
                theme = "business"  # 默认商务主题
                user_input_lower = user_input.lower()
                if (
                    "tech" in user_input_lower
                    or "技术" in user_input_lower
                    or "科技" in user_input_lower
                ):
                    theme = "tech"
                elif (
                    "creative" in user_input_lower
                    or "创意" in user_input_lower
                    or "艺术" in user_input_lower
                ):
                    theme = "creative"
                elif (
                    "simple" in user_input_lower
                    or "minimal" in user_input_lower
                    or "极简" in user_input_lower
                ):
                    theme = "minimal"

                _report("正在生成PPT...", f"主题: {theme} (自动配图)")

                ppt_gen = PPTGenerator(theme=theme)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = (
                    f"{ppt_outline.get('title', 'Presentation')}_{timestamp}.pptx"
                )
                # Max length for filename safety
                if len(filename) > 50:
                    filename = f"Presentation_{timestamp}.pptx"

                ppt_path = os.path.join(settings_manager.documents_dir, filename)
                os.makedirs(settings_manager.documents_dir, exist_ok=True)

                def _ppt_progress_wrapper(c, t, st, ty):
                    try:
                        _report(
                            f"正在生成PPT ({c}/{t})", f"页面: {st[:10]}... [{ty}]"
                        )
                    except Exception:
                        import logging; logging.getLogger(__name__).warning("Silenced exception caught", exc_info=True)

                ppt_gen.generate_from_outline(
                    title=ppt_outline.get("title", "演示"),
                    outline=ppt_outline.get("slides", []),
                    output_path=ppt_path,
                    enable_ai_images=True,
                    progress_callback=_ppt_progress_wrapper,
                )

                rel_path = os.path.relpath(ppt_path, WORKSPACE_DIR).replace(
                    "\\", "/"
                )
                saved_files.append(rel_path)
                file_type = "pptx"
                _report("PPT生成完成", f"已保存到: {rel_path}")

            except Exception as ppt_err:
                _app_logger.warning(f"[FILE_GEN] ⚠️ PPT生成失败: {ppt_err}")
                _report("PPT生成失败，回退到Word...", f"错误: {str(ppt_err)[:50]}")
                # PPT失败时回退到Word
                from web.document_generator import save_docx

                saved_docx = save_docx(
                    text_out, title=title, output_dir=settings_manager.documents_dir
                )
                rel_path = os.path.relpath(saved_docx, WORKSPACE_DIR).replace(
                    "\\", "/"
                )
                saved_files.append(rel_path)
                file_type = "docx"
        else:
            # 生成Excel或Word
            _report("正在处理内容...", "解析文档结构")
            table_rows = _extract_markdown_table(text_out)
            if prefer_excel and not table_rows:
                # 第一次未生成合格表格 → 生成修正Prompt重试一次
                fix_prompt = (
                    "请只输出一个 Markdown 表格，不要输出其他说明。\n"
                    "表格必须包含以下列：时间、价格、变化、来源。\n"
                    "每行数据一行，格式严格。\n\n"
                    f"用户需求: {context['original_input']}\n\n"
                    f"可用数据:\n{previous_data}\n"
                )
                text_out_retry = _generate_text(fix_prompt)
                if text_out_retry:
                    text_out = _clean_filegen_text(text_out_retry)
                    table_rows = _extract_markdown_table(text_out)

            if prefer_excel and table_rows:
                _report("正在生成Excel...", f"写入 {len(table_rows)} 行数据")
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import (
                        Alignment,
                        Border,
                        Font,
                        PatternFill,
                        Side,
                    )
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    ws = wb.active
                    ws.title = title[:31] if title else "Sheet1"

                    # 写入数据（清洗每个单元格内的 Markdown 符号）
                    try:
                        from web.file_quality_checker import (
                            strip_markdown_from_cell,
                        )

                        _strip_cell = strip_markdown_from_cell
                    except Exception:
                        _strip_cell = lambda x: x
                    for row in table_rows:
                        ws.append(
                            [
                                _strip_cell(str(c)) if isinstance(c, str) else c
                                for c in row
                            ]
                        )

                    # --- 样式美化 ---
                    header_font = Font(
                        name="Microsoft YaHei", size=11, bold=True, color="FFFFFF"
                    )
                    header_fill = PatternFill(
                        start_color="4472C4", end_color="4472C4", fill_type="solid"
                    )
                    data_font = Font(name="Microsoft YaHei", size=10)
                    thin_border = Border(
                        left=Side(style="thin", color="D9D9D9"),
                        right=Side(style="thin", color="D9D9D9"),
                        top=Side(style="thin", color="D9D9D9"),
                        bottom=Side(style="thin", color="D9D9D9"),
                    )
                    alt_fill = PatternFill(
                        start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"
                    )
                    center_align = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    left_align = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

                    max_row = ws.max_row
                    max_col = ws.max_column

                    for col_idx in range(1, max_col + 1):
                        # 表头样式
                        cell = ws.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border

                        # 数据行样式
                        for row_idx in range(2, max_row + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.font = data_font
                            cell.alignment = left_align
                            cell.border = thin_border
                            # 隔行变色
                            if row_idx % 2 == 0:
                                cell.fill = alt_fill

                        # 自动列宽
                        max_len = 0
                        for row_idx in range(1, max_row + 1):
                            val = ws.cell(row=row_idx, column=col_idx).value
                            if val:
                                # CJK 字符算2个字符宽
                                vlen = sum(
                                    2 if ord(c) > 127 else 1 for c in str(val)
                                )
                                max_len = max(max_len, vlen)
                        ws.column_dimensions[get_column_letter(col_idx)].width = (
                            min(max_len + 4, 40)
                        )

                    # 冻结首行
                    ws.freeze_panes = "A2"

                    filename = (
                        f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    )
                    excel_path = os.path.join(
                        settings_manager.documents_dir, filename
                    )
                    os.makedirs(settings_manager.documents_dir, exist_ok=True)
                    wb.save(excel_path)
                    rel_path = os.path.relpath(excel_path, WORKSPACE_DIR).replace(
                        "\\", "/"
                    )
                    saved_files.append(rel_path)
                    file_type = "xlsx"
                    _report("Excel生成完成", f"已保存到: {rel_path}")
                except Exception as excel_err:
                    excel_error = str(excel_err)
                    _app_logger.warning(
                        f"[FILE_GEN] ⚠️ Excel保存失败: {excel_error}"
                    )
                    _report(
                        "Excel保存失败，回退到Word...", f"错误: {excel_error[:50]}"
                    )

            # 保存为 DOCX（无表格或Excel失败时回退）
            if not saved_files:
                # ── 导出检查层（Check Layer）：质量检查 + Markdown 符号去除（永久特性）──
                try:
                    from web.file_quality_checker import FileQualityGate

                    _dqg = FileQualityGate.check_and_fix_for_export(
                        text_out,
                        target_format="word",
                        user_request=user_input,
                        progress_callback=_report,
                    )
                    text_out = _dqg["text"]
                    if _dqg.get("issues"):
                        _app_logger.debug(
                            f"[FILE_GEN] 🔍 检查层: {', '.join(_dqg['issues'][:3])}"
                        )
                except Exception as _dqge:
                    _app_logger.warning(f"[FILE_GEN] ⚠️ 导出检查层异常: {_dqge}")

                _report("正在生成Word文档...", "转换为 DOCX")
                from web.document_generator import save_docx, save_pdf

                saved_docx = save_docx(
                    text_out, title=title, output_dir=settings_manager.documents_dir
                )
                rel_path = os.path.relpath(saved_docx, WORKSPACE_DIR).replace(
                    "\\", "/"
                )
                saved_files.append(rel_path)
                file_type = "docx"
                _report("Word文档生成完成", f"已保存到: {rel_path}")

                # 如用户明确需要 PDF，也同时保存
                if prefer_pdf:
                    try:
                        _report("正在生成PDF...", "转换为 PDF")
                        saved_pdf = save_pdf(
                            text_out,
                            title=title,
                            output_dir=settings_manager.documents_dir,
                        )
                        pdf_rel = os.path.relpath(saved_pdf, WORKSPACE_DIR).replace(
                            "\\", "/"
                        )
                        saved_files.append(pdf_rel)
                        _report("PDF生成完成", f"已保存到: {pdf_rel}")
                    except Exception as pdf_err:
                        _app_logger.warning(f"[FILE_GEN] ⚠️ PDF保存失败: {pdf_err}")
                        _report("PDF生成失败", str(pdf_err)[:50])

        return {
            "success": True,
            "output": f"已生成{file_type.upper()}文档: {', '.join([os.path.basename(p) for p in saved_files])}"
            + (f" (Excel失败: {excel_error})" if excel_error else ""),
            "content": text_out,
            "file_type": file_type or "docx",
            "saved_files": saved_files,
            "model_id": model_id,
        }
    except Exception as e:
        return {"success": False, "output": "", "error": str(e)}
