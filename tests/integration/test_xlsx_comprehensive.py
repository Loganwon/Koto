# -*- coding: utf-8 -*-
"""
Comprehensive Excel / XLSX integration tests.

Covers every layer of the Excel pipeline that lacked dedicated tests:

  Layer 1  parse_xlsx  — openpyxl → Univer IWorkbookData snapshot
     - Cell value types (string, number, boolean, None)
     - Style extraction (bold, italic, font-size, font-color, bg-fill, alignment)
     - Merged-cell ranges
     - Multi-sheet workbooks
     - Required Univer envelope fields (appVersion, locale, styles, resources)
     - Min rowCount/columnCount bounds (30 / 10)
     - Empty workbook edge case

  Layer 2  export_xlsx  — Univer IWorkbookData → openpyxl .xlsx bytes
     - Cell values survive round-trip
     - Styles survive round-trip  (bold, color, bg, alignment)
     - Merged cells survive round-trip
     - Multi-sheet export
     - Empty / invalid data graceful handling
     - Image embedding from base64

  Layer 3  _openpyxl_cell_to_univer  — single-cell conversion unit tests
     - Value mapping (str→STRING, int/float→NUMBER, bool→BOOLEAN, None→skip)
     - Style sub-dict construction

  Layer 4  HTTP endpoints (open_file, open_file_by_path, auto_save, save_file)
     - Upload a real multi-sheet XLSX → parse → verify structure
     - Round-trip fidelity: create → open → edit → save → re-open → verify cell values
     - Multi-sheet round-trip
     - CJK / Unicode text in cells
     - Date cell handling (data_only=True → value, not formula)

  Layer 5  Edge cases
     - Formula cells (data_only mode returns cached value)
     - Very wide / very tall sheets
     - Sheet names with special characters
     - Workbook with zero data rows (all-empty sheets)
"""

from __future__ import annotations

import io
import json
import os
from pathlib import Path
from typing import Any

import pytest

# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build a real .xlsx in memory.  sheets = {"SheetName": [[row1], [row2], ...]}"""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_styled_xlsx() -> bytes:
    """Build a real .xlsx with styles, merges, and multiple sheets."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: styled cells
    ws1 = wb.create_sheet(title="Styled")
    c = ws1.cell(row=1, column=1, value="Bold Header")
    c.font = Font(bold=True, size=14, color="FF0000")
    c.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="top")

    ws1.cell(row=1, column=2, value="Italic").font = Font(italic=True)
    ws1.cell(row=2, column=1, value=42)
    ws1.cell(row=2, column=2, value=True)
    ws1.cell(row=3, column=1, value="")  # empty string

    # Merged range B3:C4
    ws1.merge_cells("B3:C4")
    ws1.cell(row=3, column=2, value="Merged")

    # Sheet 2: CJK data
    ws2 = wb.create_sheet(title="数据表")
    ws2.cell(row=1, column=1, value="产品")
    ws2.cell(row=1, column=2, value="销量")
    ws2.cell(row=2, column=1, value="苹果")
    ws2.cell(row=2, column=2, value=100)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _resolve_cell_style(data: dict, sheet_idx_or_id, row: int, col: int) -> dict:
    """Return the resolved style dict for a cell from parse_xlsx output.

    parse_xlsx stores cell["s"] as a string style-ID keying into
    data["styles"].  This helper resolves the lookup so tests can
    assert style properties directly.
    """
    if isinstance(sheet_idx_or_id, int):
        sid_key = data["sheetOrder"][sheet_idx_or_id]
    else:
        sid_key = sheet_idx_or_id
    cd = data["sheets"][sid_key]["cellData"]
    # cellData keys can be int (parse_xlsx direct) or str (API JSON response)
    rk = str(row) if str(row) in cd else row
    ck = (
        str(col)
        if (rk in cd and isinstance(cd[rk], dict) and str(col) in cd[rk])
        else col
    )
    cell = cd[rk][ck]
    s = cell.get("s")
    if isinstance(s, str):
        return data.get("styles", {}).get(s, {})
    return s or {}


def _make_formula_xlsx() -> bytes:
    """Build xlsx with a formula cell — data_only should return cached value."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"
    ws.cell(row=1, column=1, value=10)
    ws.cell(row=1, column=2, value=20)
    ws.cell(row=1, column=3).value = "=A1+B1"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# Layer 1: parse_xlsx unit tests
# ═══════════════════════════════════════════════════════════════════════════════


class TestParseXlsx:
    """Direct tests of app.core.file.file_parser.parse_xlsx."""

    def _parse(self, xlsx_bytes: bytes, tmp_path: Path) -> dict:
        from app.core.file.file_parser import parse_xlsx

        p = tmp_path / "test.xlsx"
        p.write_bytes(xlsx_bytes)
        return parse_xlsx(str(p))

    # ── Envelope fields ──────────────────────────────────────────────────

    def test_envelope_has_required_keys(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [["a"]]}), tmp_path)
        for key in (
            "id",
            "name",
            "appVersion",
            "locale",
            "sheetOrder",
            "sheets",
            "styles",
            "resources",
        ):
            assert key in data, f"Missing required envelope key: {key}"

    def test_app_version_and_locale_present(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [["a"]]}), tmp_path)
        assert data["appVersion"] == "0.5.0"
        assert data["locale"] == "zh-CN"

    def test_styles_is_dict_and_resources_is_list(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [["a"]]}), tmp_path)
        assert isinstance(data["styles"], dict)
        assert isinstance(data["resources"], list)

    # ── Sheet structure ──────────────────────────────────────────────────

    def test_single_sheet_structure(self, tmp_path):
        data = self._parse(_make_xlsx({"Sheet1": [["A", "B"], [1, 2]]}), tmp_path)
        assert len(data["sheetOrder"]) == 1
        sid = data["sheetOrder"][0]
        sheet = data["sheets"][sid]
        assert sheet["name"] == "Sheet1"
        assert "cellData" in sheet
        assert "mergeData" in sheet

    def test_multi_sheet_order(self, tmp_path):
        data = self._parse(
            _make_xlsx({"Alpha": [["a"]], "Beta": [["b"]], "Gamma": [["c"]]}),
            tmp_path,
        )
        assert len(data["sheetOrder"]) == 3
        names = [data["sheets"][sid]["name"] for sid in data["sheetOrder"]]
        assert names == ["Alpha", "Beta", "Gamma"]

    def test_min_row_and_column_counts(self, tmp_path):
        """rowCount >= 30 and columnCount >= 10, even for tiny sheets."""
        data = self._parse(_make_xlsx({"S1": [["x"]]}), tmp_path)
        sheet = data["sheets"][data["sheetOrder"][0]]
        assert sheet["rowCount"] >= 30
        assert sheet["columnCount"] >= 10

    def test_large_sheet_preserves_actual_dimensions(self, tmp_path):
        """A 50-row, 15-col sheet should use actual dimensions."""
        rows = [[f"r{r}c{c}" for c in range(15)] for r in range(50)]
        data = self._parse(_make_xlsx({"Big": rows}), tmp_path)
        sheet = data["sheets"][data["sheetOrder"][0]]
        assert sheet["rowCount"] >= 50
        assert sheet["columnCount"] >= 15

    # ── Cell value types ─────────────────────────────────────────────────

    def test_string_cell_type(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [["hello"]]}), tmp_path)
        cell = data["sheets"][data["sheetOrder"][0]]["cellData"][0][0]
        assert cell["v"] == "hello"
        assert cell["t"] == 1  # STRING

    def test_number_cell_type_int(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [[42]]}), tmp_path)
        cell = data["sheets"][data["sheetOrder"][0]]["cellData"][0][0]
        assert cell["v"] == 42
        assert cell["t"] == 2  # NUMBER

    def test_number_cell_type_float(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [[3.14]]}), tmp_path)
        cell = data["sheets"][data["sheetOrder"][0]]["cellData"][0][0]
        assert cell["v"] == pytest.approx(3.14)
        assert cell["t"] == 2

    def test_boolean_cell_type(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [[True]]}), tmp_path)
        cell = data["sheets"][data["sheetOrder"][0]]["cellData"][0][0]
        assert cell["v"] == 1  # bool → int
        assert cell["t"] == 3  # BOOLEAN

    def test_none_cell_is_absent(self, tmp_path):
        """None cells should not appear in cellData."""
        data = self._parse(_make_xlsx({"S1": [[None, "x"]]}), tmp_path)
        cd = data["sheets"][data["sheetOrder"][0]]["cellData"]
        # row 0, col 0 should not exist; col 1 should
        assert 0 not in cd.get(0, {})
        assert cd[0][1]["v"] == "x"

    def test_cjk_text_preserved(self, tmp_path):
        data = self._parse(_make_xlsx({"表": [["你好世界"]]}), tmp_path)
        cell = data["sheets"][data["sheetOrder"][0]]["cellData"][0][0]
        assert cell["v"] == "你好世界"

    # ── Style extraction ─────────────────────────────────────────────────

    def test_bold_style_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 0)
        assert style.get("bl") == 1

    def test_italic_style_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 1)
        assert style.get("it") == 1

    def test_font_size_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 0)
        assert style.get("fs") == 14

    def test_font_color_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 0)
        cl = style.get("cl", {})
        assert cl.get("rgb", "").upper().endswith("FF0000")

    def test_background_fill_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 0)
        bg = style.get("bg", {})
        assert bg.get("rgb", "").upper().endswith("FFFF00")

    def test_horizontal_alignment_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 0)
        # center → 2
        assert style.get("ht") == 2

    def test_vertical_alignment_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(data, 0, 0, 0)
        # top → 1
        assert style.get("vt") == 1

    # ── Merged cells ─────────────────────────────────────────────────────

    def test_merge_data_extracted(self, tmp_path):
        data = self._parse(_make_styled_xlsx(), tmp_path)
        merges = data["sheets"][data["sheetOrder"][0]]["mergeData"]
        assert len(merges) >= 1
        m = merges[0]
        # B3:C4 → (row=2, col=1) to (row=3, col=2) (0-indexed)
        assert m == {
            "startRow": 2,
            "startColumn": 1,
            "endRow": 3,
            "endColumn": 2,
        }

    def test_no_merges_yields_empty_list(self, tmp_path):
        data = self._parse(_make_xlsx({"S1": [["a"]]}), tmp_path)
        merges = data["sheets"][data["sheetOrder"][0]]["mergeData"]
        assert merges == []

    # ── Formula cells (data_only) ────────────────────────────────────────

    def test_formula_cell_returns_cached_value_or_none(self, tmp_path):
        """data_only=True returns the cached value (None if never calculated)."""
        data = self._parse(_make_formula_xlsx(), tmp_path)
        cd = data["sheets"][data["sheetOrder"][0]]["cellData"]
        # A1=10, B1=20 should exist
        assert cd[0][0]["v"] == 10
        assert cd[0][1]["v"] == 20
        # C1 has formula =A1+B1; data_only gives cached value (None for
        # newly-created workbooks that have never been opened in Excel)
        # So C1 might be absent or have v=None → we just make sure no crash
        # and the non-formula cells are correct.

    # ── Empty workbook ───────────────────────────────────────────────────

    def test_empty_sheets_parse_without_error(self, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "Empty"
        buf = io.BytesIO()
        wb.save(buf)
        data = self._parse(buf.getvalue(), tmp_path)
        assert len(data["sheetOrder"]) == 1
        cd = data["sheets"][data["sheetOrder"][0]]["cellData"]
        assert cd == {}  # no cells

    # ── Special sheet names ──────────────────────────────────────────────

    def test_special_sheet_name_preserved(self, tmp_path):
        data = self._parse(_make_xlsx({"Q1 报表 (2026)": [["x"]]}), tmp_path)
        name = data["sheets"][data["sheetOrder"][0]]["name"]
        assert name == "Q1 报表 (2026)"


# ═══════════════════════════════════════════════════════════════════════════════
# Layer 2: export_xlsx unit tests
# ═══════════════════════════════════════════════════════════════════════════════


class TestExportXlsx:
    """Direct tests of app.core.file.file_parser.export_xlsx."""

    def _export_and_reload(self, wb_data, images=None):
        """Export to bytes, then reload with openpyxl for inspection."""
        import openpyxl

        from app.core.file.file_parser import export_xlsx

        raw = export_xlsx(wb_data, images)
        assert raw[:2] == b"PK", "Output must be a valid ZIP (xlsx)"
        return openpyxl.load_workbook(io.BytesIO(raw))

    def _simple_univer(self, cells=None, name="Sheet1", merges=None):
        """Build a minimal Univer IWorkbookData dict."""
        sheet = {
            "id": "s1",
            "name": name,
            "rowCount": 30,
            "columnCount": 10,
            "cellData": cells or {},
        }
        if merges:
            sheet["mergeData"] = merges
        return {
            "sheetOrder": ["s1"],
            "sheets": {"s1": sheet},
        }

    # ── Basic cell values ────────────────────────────────────────────────

    def test_string_value_exported(self):
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": "Hello", "t": 1}}})
        )
        assert wb.active.cell(1, 1).value == "Hello"

    def test_number_value_exported(self):
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": 99.5, "t": 2}}})
        )
        assert wb.active.cell(1, 1).value == 99.5

    def test_boolean_value_exported(self):
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": 1, "t": 3}}})
        )
        assert wb.active.cell(1, 1).value == 1

    # ── Multi-sheet export ───────────────────────────────────────────────

    def test_multi_sheet_names(self):
        data = {
            "sheetOrder": ["s1", "s2"],
            "sheets": {
                "s1": {
                    "id": "s1",
                    "name": "First",
                    "cellData": {},
                    "rowCount": 30,
                    "columnCount": 10,
                },
                "s2": {
                    "id": "s2",
                    "name": "Second",
                    "cellData": {"0": {"0": {"v": "B"}}},
                    "rowCount": 30,
                    "columnCount": 10,
                },
            },
        }
        wb = self._export_and_reload(data)
        assert wb.sheetnames == ["First", "Second"]
        assert wb["Second"].cell(1, 1).value == "B"

    # ── Style export ─────────────────────────────────────────────────────

    def test_bold_style_exported(self):
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": "B", "s": {"bl": 1}}}})
        )
        assert wb.active.cell(1, 1).font.bold is True

    def test_italic_style_exported(self):
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": "I", "s": {"it": 1}}}})
        )
        assert wb.active.cell(1, 1).font.italic is True

    def test_font_size_exported(self):
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": "Big", "s": {"fs": 20}}}})
        )
        assert wb.active.cell(1, 1).font.size == 20

    def test_font_color_exported(self):
        wb = self._export_and_reload(
            self._simple_univer(
                {"0": {"0": {"v": "Red", "s": {"cl": {"rgb": "#FF0000"}}}}}
            )
        )
        color = wb.active.cell(1, 1).font.color
        assert color is not None
        assert "FF0000" in (color.rgb or "").upper()

    def test_background_fill_exported(self):
        wb = self._export_and_reload(
            self._simple_univer(
                {"0": {"0": {"v": "Y", "s": {"bg": {"rgb": "#FFFF00"}}}}}
            )
        )
        fill = wb.active.cell(1, 1).fill
        assert fill.fgColor is not None
        assert "FFFF00" in (fill.fgColor.rgb or "").upper()

    def test_alignment_exported(self):
        wb = self._export_and_reload(
            self._simple_univer(
                {"0": {"0": {"v": "C", "s": {"ht": 2, "vt": 1}}}}  # center, top
            )
        )
        ali = wb.active.cell(1, 1).alignment
        assert ali.horizontal == "center"
        assert ali.vertical == "top"

    # ── Merged cells export ──────────────────────────────────────────────

    def test_merge_data_exported(self):
        merges = [{"startRow": 0, "startColumn": 0, "endRow": 1, "endColumn": 1}]
        wb = self._export_and_reload(
            self._simple_univer(
                {"0": {"0": {"v": "Merged"}}},
                merges=merges,
            )
        )
        merged = list(wb.active.merged_cells.ranges)
        assert len(merged) == 1
        mr = merged[0]
        assert (mr.min_row, mr.min_col, mr.max_row, mr.max_col) == (1, 1, 2, 2)

    # ── Empty / edge cases ───────────────────────────────────────────────

    def test_empty_univer_data(self):
        """Empty sheetOrder produces a workbook with a fallback sheet."""
        from app.core.file.file_parser import export_xlsx

        raw = export_xlsx({"sheetOrder": [], "sheets": {}})
        assert len(raw) > 0
        assert raw[:2] == b"PK"  # valid xlsx/zip

    def test_non_list_non_dict_produces_empty(self):
        from app.core.file.file_parser import export_xlsx

        raw = export_xlsx("not a dict or list")
        assert len(raw) > 0  # should not crash
        assert raw[:2] == b"PK"

    def test_cell_without_v_key_skipped(self):
        """A cell dict with no 'v' key should still be written (for style-only cells)."""
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"s": {"bl": 1}}}})
        )
        cell = wb.active.cell(1, 1)
        # Value is None but bold style should still be set
        assert cell.font.bold is True

    # ── Image embedding ──────────────────────────────────────────────────

    def test_image_embedding_does_not_crash(self):
        """Base64 image embedding should not crash (even if we can't verify image in test)."""
        # 1x1 red PNG
        import base64

        png_1x1 = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01"
            b"\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00"
            b"\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00"
            b"\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        b64 = base64.b64encode(png_1x1).decode()
        images = [{"src": f"data:image/png;base64,{b64}"}]
        wb = self._export_and_reload(
            self._simple_univer({"0": {"0": {"v": "img"}}}),
            images=images,
        )
        assert wb.active.cell(1, 1).value == "img"

    def test_bad_image_src_silently_skipped(self):
        images = [{"src": "not-a-data-uri"}, {"src": ""}, {}]
        from app.core.file.file_parser import export_xlsx

        raw = export_xlsx(
            self._simple_univer({"0": {"0": {"v": "ok"}}}),
            images,
        )
        assert len(raw) > 0  # no crash


# ═══════════════════════════════════════════════════════════════════════════════
# Layer 3: _openpyxl_cell_to_univer unit tests
# ═══════════════════════════════════════════════════════════════════════════════


class TestCellToUniver:
    """Test the low-level cell conversion helper."""

    def _make_cell(
        self,
        value,
        bold=False,
        italic=False,
        font_size=None,
        font_color_rgb=None,
        fill_rgb=None,
        h_align=None,
        v_align=None,
    ):
        """Create a mock openpyxl cell with the given properties."""
        from unittest.mock import MagicMock

        cell = MagicMock()
        cell.value = value
        cell.row = 1
        cell.column = 1

        font = MagicMock()
        font.bold = bold
        font.italic = italic
        font.size = font_size

        if font_color_rgb:
            color = MagicMock()
            color.type = "rgb"
            color.rgb = font_color_rgb
            font.color = color
        else:
            font.color = None
        cell.font = font

        fill = MagicMock()
        if fill_rgb:
            fill.fill_type = "solid"
            fg = MagicMock()
            fg.type = "rgb"
            fg.rgb = fill_rgb
            fill.fgColor = fg
        else:
            fill.fill_type = None
            fill.fgColor = None
        cell.fill = fill

        alignment = MagicMock()
        alignment.horizontal = h_align
        alignment.vertical = v_align
        cell.alignment = alignment

        return cell

    def test_none_value_returns_none(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        cell = self._make_cell(None)
        assert _openpyxl_cell_to_univer(cell) is None

    def test_string_value(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("hello"))
        assert result["v"] == "hello"
        assert result["t"] == 1

    def test_integer_value(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell(42))
        assert result["v"] == 42
        assert result["t"] == 2

    def test_float_value(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell(3.14))
        assert result["v"] == pytest.approx(3.14)
        assert result["t"] == 2

    def test_boolean_value(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell(True))
        assert result["v"] == 1
        assert result["t"] == 3

    def test_bold_style(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", bold=True))
        assert result["s"]["bl"] == 1

    def test_italic_style(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", italic=True))
        assert result["s"]["it"] == 1

    def test_font_size_style(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", font_size=16))
        assert result["s"]["fs"] == 16

    def test_font_color_style(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(
            self._make_cell("x", font_color_rgb="FFFF0000")
        )
        assert result["s"]["cl"]["rgb"] == "#FF0000"

    def test_fill_color_style(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", fill_rgb="FF00FF00"))
        assert result["s"]["bg"]["rgb"] == "#00FF00"

    def test_alignment_horizontal(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", h_align="center"))
        assert result["s"]["ht"] == 2

    def test_alignment_vertical(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", v_align="bottom"))
        assert result["s"]["vt"] == 3

    def test_no_style_means_no_s_key(self):
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("plain"))
        assert "s" not in result

    def test_default_black_color_ignored(self):
        """Standard black (00000000 or FF000000) should not create a cl entry."""
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(
            self._make_cell("x", font_color_rgb="FF000000")
        )
        assert "cl" not in result.get("s", {})

    def test_white_fill_ignored(self):
        """Standard white fill (FFFFFFFF) should not create a bg entry."""
        from app.core.file.file_parser import _openpyxl_cell_to_univer

        result = _openpyxl_cell_to_univer(self._make_cell("x", fill_rgb="FFFFFFFF"))
        assert "bg" not in result.get("s", {})


# ═══════════════════════════════════════════════════════════════════════════════
# Layer 4: Full round-trip (parse → export → re-parse)
# ═══════════════════════════════════════════════════════════════════════════════


class TestRoundTrip:
    """Parse a real xlsx, export via export_xlsx, re-parse, and verify fidelity."""

    def _round_trip(self, xlsx_bytes: bytes, tmp_path: Path) -> tuple[dict, dict]:
        """Returns (original_data, re_parsed_data)."""
        from app.core.file.file_parser import export_xlsx, parse_xlsx

        # Parse original
        p1 = tmp_path / "original.xlsx"
        p1.write_bytes(xlsx_bytes)
        data1 = parse_xlsx(str(p1))

        # Export
        raw = export_xlsx(data1)

        # Re-parse
        p2 = tmp_path / "exported.xlsx"
        p2.write_bytes(raw)
        data2 = parse_xlsx(str(p2))

        return data1, data2

    def test_cell_values_survive_round_trip(self, tmp_path):
        rows = [["Name", "Age", "Score"], ["Alice", 30, 95.5], ["Bob", 25, 88.0]]
        d1, d2 = self._round_trip(_make_xlsx({"Data": rows}), tmp_path)

        cd1 = d1["sheets"][d1["sheetOrder"][0]]["cellData"]
        cd2 = d2["sheets"][d2["sheetOrder"][0]]["cellData"]

        for rk, row in cd1.items():
            for ck, cell in row.items():
                assert (
                    cd2[rk][ck]["v"] == cell["v"]
                ), f"Cell ({rk},{ck}) value mismatch: {cell['v']} → {cd2[rk][ck]['v']}"

    def test_multi_sheet_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(
            _make_xlsx({"Sales": [["Q1", 100]], "Costs": [["Q1", 50]]}),
            tmp_path,
        )
        names1 = [d1["sheets"][sid]["name"] for sid in d1["sheetOrder"]]
        names2 = [d2["sheets"][sid]["name"] for sid in d2["sheetOrder"]]
        assert names1 == names2

    def test_cjk_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(
            _make_xlsx({"中文表": [["名称", "数量"], ["苹果", 100]]}),
            tmp_path,
        )
        cd = d2["sheets"][d2["sheetOrder"][0]]["cellData"]
        assert cd[0][0]["v"] == "名称"
        assert cd[1][0]["v"] == "苹果"

    def test_bold_style_survives_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(d2, 0, 0, 0)
        assert style.get("bl") == 1, "Bold style lost in round-trip"

    def test_merge_data_survives_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(_make_styled_xlsx(), tmp_path)
        m1 = d1["sheets"][d1["sheetOrder"][0]]["mergeData"]
        m2 = d2["sheets"][d2["sheetOrder"][0]]["mergeData"]
        assert len(m2) == len(m1), f"Merge count mismatch: {len(m1)} → {len(m2)}"
        # The merge range coordinates should match
        assert m2[0]["startRow"] == m1[0]["startRow"]
        assert m2[0]["startColumn"] == m1[0]["startColumn"]
        assert m2[0]["endRow"] == m1[0]["endRow"]
        assert m2[0]["endColumn"] == m1[0]["endColumn"]

    def test_font_color_survives_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(d2, 0, 0, 0)
        cl = style.get("cl", {})
        assert "FF0000" in cl.get("rgb", "").upper(), "Font color lost in round-trip"

    def test_background_fill_survives_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(d2, 0, 0, 0)
        bg = style.get("bg", {})
        assert (
            "FFFF00" in bg.get("rgb", "").upper()
        ), "Background fill lost in round-trip"

    def test_alignment_survives_round_trip(self, tmp_path):
        d1, d2 = self._round_trip(_make_styled_xlsx(), tmp_path)
        style = _resolve_cell_style(d2, 0, 0, 0)
        assert style.get("ht") == 2, "Horizontal alignment lost"
        assert style.get("vt") == 1, "Vertical alignment lost"


# ═══════════════════════════════════════════════════════════════════════════════
# Layer 5: HTTP endpoint tests (Flask test client)
# ═══════════════════════════════════════════════════════════════════════════════


@pytest.fixture(scope="module")
def xlsx_client(tmp_path_factory):
    """Flask test client with isolated workspace, focused on xlsx tests."""
    os.environ.setdefault("KOTO_AUTH_ENABLED", "false")

    tmp_root = tmp_path_factory.mktemp("xlsx_test")
    tmp_dir = tmp_root / "tmp"
    workspace_dir = tmp_root / "workspace"
    tmp_dir.mkdir(parents=True)
    workspace_dir.mkdir(parents=True)

    import web.blueprints.workspace_assistant as _wa
    import web.shared as _shared

    original_tmp = _wa._TMP_DIR
    original_ws = getattr(_shared, "WORKSPACE_DIR", None)
    _wa._TMP_DIR = tmp_dir
    _shared.WORKSPACE_DIR = str(workspace_dir)

    from flask import Flask

    from web.blueprints.workspace_assistant import workspace_assistant_bp

    app = Flask(__name__)
    app.register_blueprint(workspace_assistant_bp)
    app.config["TESTING"] = True

    with app.test_client() as client:
        yield client, tmp_dir, workspace_dir

    _wa._TMP_DIR = original_tmp
    if original_ws is not None:
        _shared.WORKSPACE_DIR = original_ws


class TestXlsxUploadEndpoint:
    """POST /api/v1/workspace/open_file — upload a real xlsx."""

    def test_upload_real_xlsx_returns_univer_data(self, xlsx_client):
        client, _, _ = xlsx_client
        xlsx_bytes = _make_xlsx({"Sales": [["Item", "Qty"], ["Apple", 100]]})
        resp = client.post(
            "/api/v1/workspace/open_file",
            data={"file": (io.BytesIO(xlsx_bytes), "sales.xlsx")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        body = resp.get_json()
        assert body["file_type"] == "xlsx"
        assert body["file_name"] == "sales.xlsx"
        data = body["data"]
        assert "sheetOrder" in data
        assert "sheets" in data
        sid = data["sheetOrder"][0]
        cd = data["sheets"][sid]["cellData"]
        assert cd["0"]["0"]["v"] == "Item"
        assert cd["1"]["1"]["v"] == 100

    def test_upload_multi_sheet_xlsx(self, xlsx_client):
        client, _, _ = xlsx_client
        xlsx_bytes = _make_xlsx({"A": [["a1"]], "B": [["b1"]], "C": [["c1"]]})
        resp = client.post(
            "/api/v1/workspace/open_file",
            data={"file": (io.BytesIO(xlsx_bytes), "multi.xlsx")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        data = resp.get_json()["data"]
        assert len(data["sheetOrder"]) == 3

    def test_upload_styled_xlsx_preserves_styles(self, xlsx_client):
        client, _, _ = xlsx_client
        resp = client.post(
            "/api/v1/workspace/open_file",
            data={"file": (io.BytesIO(_make_styled_xlsx()), "styled.xlsx")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        data = resp.get_json()["data"]
        cd = data["sheets"][data["sheetOrder"][0]]["cellData"]
        style = _resolve_cell_style(data, 0, 0, 0)
        assert style.get("bl") == 1, "Bold not preserved on upload"

    def test_upload_xlsx_with_merges_preserves_merge_data(self, xlsx_client):
        client, _, _ = xlsx_client
        resp = client.post(
            "/api/v1/workspace/open_file",
            data={"file": (io.BytesIO(_make_styled_xlsx()), "merged.xlsx")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        data = resp.get_json()["data"]
        merges = data["sheets"][data["sheetOrder"][0]]["mergeData"]
        assert len(merges) >= 1


class TestXlsxAutoSaveEndpoint:
    """POST /api/v1/workspace/auto_save — xlsx round-trip through HTTP."""

    def test_auto_save_xlsx_persists_cell_values(self, xlsx_client):
        """Create → open → save with new data → reopen → values match."""
        client, _, ws = xlsx_client

        # Create
        client.post(
            "/api/v1/workspace/create_file", json={"folder": "", "name": "persist.xlsx"}
        )

        # Open
        resp = client.post(
            "/api/v1/workspace/open_file_by_path", json={"path": "persist.xlsx"}
        )
        assert resp.status_code == 200
        fid = resp.get_json()["file_id"]

        # Save with known data
        save_data = {
            "snapshot": {
                "sheetOrder": ["s1"],
                "sheets": {
                    "s1": {
                        "id": "s1",
                        "name": "Saved",
                        "rowCount": 30,
                        "columnCount": 10,
                        "cellData": {
                            "0": {"0": {"v": "Saved!", "t": 1}},
                            "1": {"0": {"v": 42, "t": 2}},
                            "2": {"0": {"v": "日本語", "t": 1}},
                        },
                    }
                },
            },
            "_images": [],
        }
        resp2 = client.post(
            "/api/v1/workspace/auto_save",
            json={
                "file_type": "xlsx",
                "file_id": fid,
                "ws_source_path": "persist.xlsx",
                "explicit": True,
                "data": save_data,
            },
        )
        assert resp2.status_code == 200
        assert resp2.get_json()["src_written"] is True

        # Reopen and verify
        resp3 = client.post(
            "/api/v1/workspace/open_file_by_path", json={"path": "persist.xlsx"}
        )
        assert resp3.status_code == 200
        cd = resp3.get_json()["data"]["sheets"][
            resp3.get_json()["data"]["sheetOrder"][0]
        ]["cellData"]
        assert cd["0"]["0"]["v"] == "Saved!"
        assert cd["1"]["0"]["v"] == 42
        assert cd["2"]["0"]["v"] == "日本語"

    def test_auto_save_xlsx_with_styles_and_merges(self, xlsx_client):
        """Styles and merges in the save payload should survive round-trip."""
        client, _, ws = xlsx_client

        client.post(
            "/api/v1/workspace/create_file",
            json={"folder": "", "name": "styled_save.xlsx"},
        )
        resp = client.post(
            "/api/v1/workspace/open_file_by_path", json={"path": "styled_save.xlsx"}
        )
        fid = resp.get_json()["file_id"]

        save_data = {
            "snapshot": {
                "sheetOrder": ["s1"],
                "sheets": {
                    "s1": {
                        "id": "s1",
                        "name": "Styled",
                        "rowCount": 30,
                        "columnCount": 10,
                        "cellData": {
                            "0": {
                                "0": {"v": "Header", "t": 1, "s": {"bl": 1, "fs": 14}}
                            },
                            "1": {"0": {"v": 100, "t": 2}},
                        },
                        "mergeData": [
                            {
                                "startRow": 2,
                                "startColumn": 0,
                                "endRow": 3,
                                "endColumn": 1,
                            }
                        ],
                    }
                },
            },
            "_images": [],
        }
        resp2 = client.post(
            "/api/v1/workspace/auto_save",
            json={
                "file_type": "xlsx",
                "file_id": fid,
                "ws_source_path": "styled_save.xlsx",
                "explicit": True,
                "data": save_data,
            },
        )
        assert resp2.status_code == 200

        # Reopen
        resp3 = client.post(
            "/api/v1/workspace/open_file_by_path", json={"path": "styled_save.xlsx"}
        )
        assert resp3.status_code == 200
        wb_data = resp3.get_json()["data"]
        sheet = wb_data["sheets"][wb_data["sheetOrder"][0]]
        # Bold style (resolved from styles registry)
        style = _resolve_cell_style(wb_data, 0, 0, 0)
        assert style.get("bl") == 1
        # Merge
        assert len(sheet["mergeData"]) >= 1

    def test_timer_auto_save_does_not_write_source(self, xlsx_client):
        """Non-explicit (timer) auto_save must NOT write to workspace source."""
        client, _, ws = xlsx_client

        client.post(
            "/api/v1/workspace/create_file", json={"folder": "", "name": "timer.xlsx"}
        )
        resp = client.post(
            "/api/v1/workspace/open_file_by_path", json={"path": "timer.xlsx"}
        )
        fid = resp.get_json()["file_id"]
        original_size = (ws / "timer.xlsx").stat().st_size

        # Timer save (explicit=False)
        save_data = {
            "snapshot": {
                "sheetOrder": ["s1"],
                "sheets": {
                    "s1": {
                        "id": "s1",
                        "name": "Timer",
                        "cellData": {"0": {"0": {"v": "CHANGED"}}},
                        "rowCount": 30,
                        "columnCount": 10,
                    }
                },
            },
        }
        resp2 = client.post(
            "/api/v1/workspace/auto_save",
            json={
                "file_type": "xlsx",
                "file_id": fid,
                "ws_source_path": "timer.xlsx",
                "explicit": False,
                "data": save_data,
            },
        )
        assert resp2.status_code == 200
        assert resp2.get_json()["src_written"] is False


class TestXlsxSaveFileEndpoint:
    """POST /api/v1/workspace/save_file — download exported xlsx."""

    def test_save_file_xlsx_returns_valid_zip(self, xlsx_client):
        client, _, _ = xlsx_client
        wb_data = {
            "snapshot": {
                "sheetOrder": ["s1"],
                "sheets": {
                    "s1": {
                        "id": "s1",
                        "name": "DL",
                        "cellData": {"0": {"0": {"v": "Download"}}},
                        "rowCount": 30,
                        "columnCount": 10,
                    }
                },
            },
            "_images": [],
        }
        resp = client.post(
            "/api/v1/workspace/save_file",
            json={
                "file_type": "xlsx",
                "data": wb_data,
                "file_name": "dl.xlsx",
            },
        )
        assert resp.status_code == 200
        assert resp.data[:2] == b"PK"

    def test_save_file_xlsx_content_is_readable(self, xlsx_client):
        """The downloaded xlsx should be openable by openpyxl."""
        import openpyxl

        client, _, _ = xlsx_client
        wb_data = {
            "snapshot": {
                "sheetOrder": ["s1"],
                "sheets": {
                    "s1": {
                        "id": "s1",
                        "name": "Verify",
                        "cellData": {"0": {"0": {"v": "Check", "t": 1}}},
                        "rowCount": 30,
                        "columnCount": 10,
                    }
                },
            },
            "_images": [],
        }
        resp = client.post(
            "/api/v1/workspace/save_file",
            json={
                "file_type": "xlsx",
                "data": wb_data,
                "file_name": "verify.xlsx",
            },
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        assert wb.active.cell(1, 1).value == "Check"

    def test_save_file_xlsx_bare_dict_format(self, xlsx_client):
        """If data is a bare IWorkbookData (no 'snapshot' wrapper), it should still work."""
        client, _, _ = xlsx_client
        bare_data = {
            "sheetOrder": ["s1"],
            "sheets": {
                "s1": {
                    "id": "s1",
                    "name": "Bare",
                    "cellData": {"0": {"0": {"v": "BareOK"}}},
                    "rowCount": 30,
                    "columnCount": 10,
                }
            },
        }
        resp = client.post(
            "/api/v1/workspace/save_file",
            json={
                "file_type": "xlsx",
                "data": bare_data,
                "file_name": "bare.xlsx",
            },
        )
        assert resp.status_code == 200
        assert resp.data[:2] == b"PK"


class TestXlsxOpenByPathExternal:
    """open_abs_file with an absolute external xlsx path."""

    def test_open_external_xlsx(self, xlsx_client, tmp_path):
        client, _, _ = xlsx_client
        ext_file = tmp_path / "external_data.xlsx"
        ext_file.write_bytes(_make_xlsx({"Ext": [["External", 999]]}))

        resp = client.post(
            "/api/v1/workspace/open_abs_file", json={"path": str(ext_file)}
        )
        assert resp.status_code == 200
        body = resp.get_json()
        assert body["file_type"] == "xlsx"
        cd = body["data"]["sheets"][body["data"]["sheetOrder"][0]]["cellData"]
        assert cd["0"]["0"]["v"] == "External"
        assert cd["0"]["1"]["v"] == 999

    def test_open_external_xlsx_with_styles(self, xlsx_client, tmp_path):
        client, _, _ = xlsx_client
        ext_file = tmp_path / "styled_ext.xlsx"
        ext_file.write_bytes(_make_styled_xlsx())

        resp = client.post(
            "/api/v1/workspace/open_abs_file", json={"path": str(ext_file)}
        )
        assert resp.status_code == 200
        wb_data = resp.get_json()["data"]
        style = _resolve_cell_style(wb_data, 0, 0, 0)
        assert style.get("bl") == 1


class TestXlsxRawFile:
    """GET /api/v1/workspace/raw/<file_id> — raw xlsx download."""

    def test_raw_xlsx_returns_valid_bytes(self, xlsx_client):
        client, _, _ = xlsx_client
        xlsx_bytes = _make_xlsx({"R": [["raw"]]})
        resp = client.post(
            "/api/v1/workspace/open_file",
            data={"file": (io.BytesIO(xlsx_bytes), "raw_test.xlsx")},
            content_type="multipart/form-data",
        )
        fid = resp.get_json()["file_id"]

        raw_resp = client.get(f"/api/v1/workspace/raw/{fid}")
        assert raw_resp.status_code == 200
        assert raw_resp.data[:2] == b"PK"
