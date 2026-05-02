"""Generate test .xlsx files for DPI scaling / Univer Sheets testing.

Run:  python tests/fixtures/generate_dpi_test_xlsx.py

Creates several xlsx files in tests/fixtures/ that exercise different
Univer rendering scenarios affected by DPI counter-zoom.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def _save(wb, name):
    path = os.path.join(OUT_DIR, name)
    wb.save(path)
    print(f"  -> {path}")


# ── 1. Basic grid with toolbar-heavy formatting ──────────────────────────
# Tests: toolbar click positioning, dropdown alignment
def make_toolbar_test():
    wb = Workbook()
    ws = wb.active
    ws.title = "工具栏测试"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["序号", "姓名", "部门", "职位", "入职日期", "薪资", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    data = [
        [1, "张三", "工程部", "高级工程师", "2023-01-15", 25000, "团队负责人"],
        [2, "李四", "市场部", "市场经理", "2022-06-20", 22000, ""],
        [3, "王五", "财务部", "财务主管", "2021-11-08", 28000, "兼任审计"],
        [4, "赵六", "人事部", "HR专员", "2024-03-01", 15000, "实习转正"],
        [5, "孙七", "工程部", "前端开发", "2023-08-12", 20000, ""],
        [6, "周八", "工程部", "后端开发", "2023-09-05", 21000, ""],
        [7, "吴九", "市场部", "品牌设计", "2024-01-10", 18000, "外包转正"],
        [8, "郑十", "产品部", "产品经理", "2022-04-18", 26000, ""],
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:
                cell.number_format = "#,##0"
            if c == 5:
                cell.number_format = "YYYY-MM-DD"

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    _save(wb, "dpi_test_toolbar.xlsx")


# ── 2. Cell editing test — merged cells, formulas, validation ────────────
# Tests: double-click inline editing, cell editor popup positioning
def make_edit_test():
    wb = Workbook()
    ws = wb.active
    ws.title = "编辑测试"

    # Merged header
    ws.merge_cells("A1:F1")
    ws["A1"] = "季度销售报表 — 双击此处编辑"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Sub-headers
    sub = ["区域", "Q1", "Q2", "Q3", "Q4", "合计"]
    for c, h in enumerate(sub, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")

    regions = ["华东", "华南", "华北", "西南", "海外"]
    import random
    random.seed(42)
    for r, region in enumerate(regions, 4):
        ws.cell(row=r, column=1, value=region)
        for c in range(2, 6):
            ws.cell(row=r, column=c, value=random.randint(50000, 200000))
            ws.cell(row=r, column=c).number_format = "#,##0"
        # SUM formula in column F
        ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
        ws.cell(row=r, column=6).number_format = "#,##0"
        ws.cell(row=r, column=6).font = Font(bold=True)

    # Long text cell for testing text overflow editing
    ws.cell(row=10, column=1, value="备注")
    ws.merge_cells("B10:F10")
    ws["B10"] = "这是一段很长的文本，用于测试在DPI缩放模式下双击编辑长文本时编辑器弹出位置是否正确。如果CSS counter-zoom工作正常，编辑框应该出现在单元格正上方，而不是偏移到右下方。"
    ws["B10"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[10].height = 60

    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 16

    _save(wb, "dpi_test_edit.xlsx")


# ── 3. Dense data for resolution / clarity test ──────────────────────────
# Tests: canvas rendering sharpness at various DPI levels
def make_resolution_test():
    wb = Workbook()
    ws = wb.active
    ws.title = "分辨率测试"

    small_font = Font(size=9)
    border = Border(
        left=Side(style="hair"),
        right=Side(style="hair"),
        top=Side(style="hair"),
        bottom=Side(style="hair"),
    )

    # 26 columns (A-Z) × 50 rows of small text
    for c in range(1, 27):
        ws.column_dimensions[get_column_letter(c)].width = 8
        header = ws.cell(row=1, column=c, value=get_column_letter(c))
        header.font = Font(bold=True, size=9)
        header.fill = PatternFill("solid", fgColor="E2EFDA")
        header.border = border

    import random
    random.seed(123)
    for r in range(2, 52):
        for c in range(1, 27):
            val = round(random.uniform(-100, 100), 2)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = small_font
            cell.number_format = "0.00"
            cell.border = border
            # Color negative values red
            if val < 0:
                cell.font = Font(size=9, color="FF0000")

    _save(wb, "dpi_test_resolution.xlsx")


# ── 4. Multi-sheet with charts data ─────────────────────────────────────
# Tests: sheet tab switching, scrolling, complex layout
def make_multisheet_test():
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "总览"
    ws1["A1"] = "多Sheet测试文件"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A3"] = "本文件包含多个Sheet页，用于测试："
    ws1["A4"] = "1. Sheet标签切换时DPI是否正常"
    ws1["A5"] = "2. 大范围滚动时画面是否清晰"
    ws1["A6"] = "3. 冻结窗格下的鼠标定位"

    # Sheet 2: Frozen panes
    ws2 = wb.create_sheet("冻结窗格")
    ws2.freeze_panes = "B3"
    for c in range(1, 16):
        ws2.cell(row=1, column=c, value=f"列{c}")
        ws2.cell(row=1, column=c).font = Font(bold=True)
    for r in range(2, 101):
        ws2.cell(row=r, column=1, value=f"行{r}")
        ws2.cell(row=r, column=1).font = Font(bold=True)
        for c in range(2, 16):
            ws2.cell(row=r, column=c, value=r * c)

    # Sheet 3: Wide columns
    ws3 = wb.create_sheet("宽列测试")
    for c in range(1, 6):
        letter = get_column_letter(c)
        ws3.column_dimensions[letter].width = 30 + c * 10
        ws3.cell(row=1, column=c, value=f"宽度={30+c*10}的列")
        ws3.cell(row=1, column=c).font = Font(bold=True, size=11)
    for r in range(2, 21):
        for c in range(1, 6):
            ws3.cell(row=r, column=c, value=f"单元格 ({r},{c}) — 测试宽列下的鼠标点击定位")

    _save(wb, "dpi_test_multisheet.xlsx")


if __name__ == "__main__":
    print("Generating DPI test xlsx files...")
    make_toolbar_test()
    make_edit_test()
    make_resolution_test()
    make_multisheet_test()
    print("Done! 4 test files created.")
