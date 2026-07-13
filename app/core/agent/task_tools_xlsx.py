# -*- coding: utf-8 -*-
# Copyright (C) 2024-2026 Koto AI. All rights reserved.
# SPDX-License-Identifier: LicenseRef-Koto-Proprietary
"""XLSX (Excel) tool implementations extracted from task_tools.py.

Functions:
    read_sheet_data          - Read spreadsheet cells as structured JSON
    inspect_workbook_structure - Inspect sheet names, headers, and formulas
    audit_financial_workbook - Audit financial data for anomalies
    write_sheet_data         - Write/update cells in a spreadsheet
"""

from __future__ import annotations

import json
import logging
from typing import Any, List, Optional

from app.core.agent.task_tools_xlsx_sheet_selection import (
    select_workbook_sheet as _select_workbook_sheet,
)
from app.core.agent.task_tools_xlsx_sheet_selection import (
    sheet_matches_statement as _sheet_matches_statement,
)


def _task_tools_helper(name: str):
    """Resolve legacy shared helpers only after the module graph is ready.

    ``task_tools`` re-exports this module's public XLSX functions for backward
    compatibility.  Importing its private helpers here at module load time made
    ``import task_tools_xlsx`` fail with a partially-initialized circular import.
    The helpers are invoked only while a tool is executing, by which point the
    parent module has completed initialization.
    """

    def call(*args: Any, **kwargs: Any) -> Any:
        from . import task_tools

        return getattr(task_tools, name)(*args, **kwargs)

    return call


_best_effort_backup = _task_tools_helper("_best_effort_backup")
_blocked_write_result = _task_tools_helper("_blocked_write_result")
_build_workbook_structure_payload = _task_tools_helper(
    "_build_workbook_structure_payload"
)
_detect_financial_series_gap_findings = _task_tools_helper(
    "_detect_financial_series_gap_findings"
)
_ensure_existing_file_writable = _task_tools_helper("_ensure_existing_file_writable")
_file_task_diff = _task_tools_helper("_file_task_diff")
_merge_warnings = _task_tools_helper("_merge_warnings")
_nonwritable_target_message = _task_tools_helper("_nonwritable_target_message")
_nonwritable_target_next_step = _task_tools_helper("_nonwritable_target_next_step")
_normalize_positive_int = _task_tools_helper("_normalize_positive_int")
_resolve_path = _task_tools_helper("_resolve_path")
_result_path = _task_tools_helper("_result_path")
_safe_resolve = _task_tools_helper("_safe_resolve")
_save_workbook_via_temp_file = _task_tools_helper("_save_workbook_via_temp_file")
_success_result = _task_tools_helper("_success_result")

_logger = logging.getLogger(__name__)


def read_sheet_data(path: str, sheet_name: str = "", max_rows: int = 500) -> str:
    """Read spreadsheet cells as structured JSON.

    Returns JSON: {"sheet": "<name>", "headers": [...], "rows": [[...], ...]}
    """
    max_rows = _normalize_positive_int(max_rows, default=500, upper=10_000)
    resolved = _resolve_path(path)
    if not resolved:
        return json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
        target_sheet, requested_sheet, sheet_warning = _select_workbook_sheet(
            wb, sheet_name
        )
        if not target_sheet:
            available_sheets = list(getattr(wb, "sheetnames", []) or [])
            wb.close()
            if requested_sheet and available_sheets:
                return json.dumps(
                    {
                        "sheet": "",
                        "headers": [],
                        "rows": [],
                        "row_count": 0,
                        "requested_sheet": requested_sheet,
                        "available_sheets": available_sheets,
                        "missing_requested_sheet": True,
                        "warning": sheet_warning,
                        "summary": f"未找到工作表 '{requested_sheet}'，已返回可用工作表列表供继续分析。",
                    },
                    ensure_ascii=False,
                )
            return json.dumps(
                {"error": sheet_warning},
                ensure_ascii=False,
            )
        ws = wb[target_sheet]
        headers: list[str] = []
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(max_row=max_rows + 1, values_only=True)):
            cells = [v if v is not None else "" for v in row]
            if i == 0:
                headers = [str(c) for c in cells]
            else:
                rows.append(cells)
        wb.close()
        payload: Dict[str, Any] = {
            "sheet": target_sheet,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
        }
        if sheet_warning:
            payload["warning"] = sheet_warning
            payload["requested_sheet"] = requested_sheet
        return json.dumps(payload, ensure_ascii=False, default=str)
    except ImportError:
        return json.dumps({"error": "openpyxl not installed"}, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


def inspect_workbook_structure(
    path: str,
    sample_rows_per_sheet: int = 6,
    max_formula_examples_per_sheet: int = 8,
) -> str:
    """Inspect workbook sheets, formulas, year headers, and external links."""
    sample_rows_per_sheet = _normalize_positive_int(
        sample_rows_per_sheet, default=6, upper=20
    )
    max_formula_examples_per_sheet = _normalize_positive_int(
        max_formula_examples_per_sheet, default=8, upper=20
    )
    resolved = _resolve_path(path)
    if not resolved:
        return json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False)

    try:
        payload = _build_workbook_structure_payload(
            resolved,
            sample_rows_per_sheet=sample_rows_per_sheet,
            max_formula_examples_per_sheet=max_formula_examples_per_sheet,
        )
        payload["path"] = _result_path(path, resolved)
        return json.dumps(payload, ensure_ascii=False, default=str)
    except ImportError:
        return json.dumps({"error": "openpyxl not installed"}, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)


def audit_financial_workbook(
    path: str,
    sample_rows_per_sheet: int = 4,
    max_formula_examples_per_sheet: int = 6,
    max_findings: int = 12,
) -> str:
    """Audit an Excel financial model for completeness, external dependencies, and row continuity gaps."""
    sample_rows_per_sheet = _normalize_positive_int(
        sample_rows_per_sheet, default=4, upper=12
    )
    max_formula_examples_per_sheet = _normalize_positive_int(
        max_formula_examples_per_sheet, default=6, upper=12
    )
    max_findings = _normalize_positive_int(max_findings, default=12, upper=30)
    resolved = _resolve_path(path)
    if not resolved:
        return json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False)

    try:
        structure = _build_workbook_structure_payload(
            resolved,
            sample_rows_per_sheet=sample_rows_per_sheet,
            max_formula_examples_per_sheet=max_formula_examples_per_sheet,
        )
        findings: List[Dict[str, Any]] = []
        for statement_key, label in (
            ("profit_and_loss", "利润表 / Income Statement"),
            ("balance_sheet", "资产负债表 / Balance Sheet"),
            ("cash_flow", "现金流量表 / Cash Flow"),
        ):
            presence = structure["statement_presence"].get(statement_key) or {}
            if presence.get("present"):
                continue
            findings.append(
                {
                    "severity": (
                        "high" if statement_key != "profit_and_loss" else "medium"
                    ),
                    "type": "missing_statement",
                    "message": f"未检测到 {label} 工作表，当前文件不是可完整勾稽的三表模型。",
                    "evidence": {"available_sheets": structure["sheet_names"]},
                }
            )

        if (
            structure["external_link_targets"]
            or structure["total_external_formula_refs"]
        ):
            findings.append(
                {
                    "severity": "high",
                    "type": "external_dependency",
                    "message": (
                        f"工作簿包含 {structure['external_link_count']} 个外部链接，"
                        "关键公式依赖外部底稿，当前文件不能视为可独立复算的完整模型。"
                    ),
                    "evidence": {
                        "external_link_targets": structure["external_link_targets"],
                        "external_formula_refs": structure[
                            "total_external_formula_refs"
                        ],
                    },
                }
            )

        findings.extend(
            _detect_financial_series_gap_findings(resolved, max_findings=max_findings)
        )
        findings = findings[:max_findings]
        high_count = sum(1 for item in findings if item.get("severity") == "high")
        medium_count = sum(1 for item in findings if item.get("severity") == "medium")
        summary_parts = ["已完成财务工作簿审计"]
        if high_count:
            summary_parts.append(f"发现 {high_count} 个高优先级问题")
        if medium_count:
            summary_parts.append(f"发现 {medium_count} 个中优先级问题")
        if len(summary_parts) == 1:
            summary_parts.append("未发现明显结构性红旗")

        payload = {
            "path": _result_path(path, resolved),
            "sheet_names": structure["sheet_names"],
            "statement_presence": structure["statement_presence"],
            "external_link_targets": structure["external_link_targets"],
            "findings": findings,
            "recommended_sheet_reads": [
                sheet["name"]
                for sheet in structure["sheets"]
                if sheet.get("formula_count")
                or _sheet_matches_statement(sheet.get("name"), "profit_and_loss")
            ][:5],
            "summary": "；".join(summary_parts) + "。",
        }
        return json.dumps(payload, ensure_ascii=False, default=str)
    except ImportError:
        return json.dumps({"error": "openpyxl not installed"}, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)


def write_sheet_data(path: str, sheet_name: str = "", updates: str = "[]") -> str:
    """Write cells to a spreadsheet.

    `updates` is a JSON array: [{"row": 1, "col": 1, "value": "..."}, ...]
    Row and col are 1-indexed.
    """
    resolved = _resolve_path(path)
    if not resolved:
        return json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False)

    try:
        cell_updates = json.loads(updates) if isinstance(updates, str) else updates
    except json.JSONDecodeError as e:
        return json.dumps({"error": f"Invalid updates JSON: {e}"}, ensure_ascii=False)

    wb = None
    count = 0
    diff_items: List[Dict[str, Any]] = []
    target_sheet = str(sheet_name or "").strip()
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        backup_warning = _best_effort_backup(resolved)
        write_warning = _ensure_existing_file_writable(resolved)

        wb = openpyxl.load_workbook(resolved)
        target = sheet_name or wb.sheetnames[0]
        if target not in wb.sheetnames:
            wb.close()
            return json.dumps(
                {"error": f"Sheet '{target}' not found"}, ensure_ascii=False
            )
        ws = wb[target]

        for u in cell_updates:
            row = int(u.get("row", 0))
            col = int(u.get("col", 0))
            value = u.get("value", "")
            if row < 1 or col < 1:
                continue
            cell = ws.cell(row=row, column=col)
            before_value = cell.value
            # Detect Excel formulas — write as formula, not literal string
            if isinstance(value, str) and value.startswith("="):
                cell.value = value
            else:
                cell.value = value
            count += 1
            cell_ref = f"{get_column_letter(col)}{row}"
            diff_items.append(
                {
                    "sheet": target,
                    "cell": cell_ref,
                    "row": row,
                    "col": col,
                    "before": before_value,
                    "after": cell.value,
                }
            )

        _save_workbook_via_temp_file(wb, resolved)
        wb.close()
        target_sheet = sheet_name or target
        return _success_result(
            _result_path(path, resolved),
            operation="write_sheet_data",
            summary=f"已写入 {count} 个单元格到工作表“{target_sheet}”",
            file_type="xlsx",
            change_type="modify",
            summary_code="WRITE_OK",
            diff=_file_task_diff("xlsx_cells", diff_items),
            cells_written=count,
            sheet=target_sheet,
            warning=_merge_warnings(backup_warning, write_warning),
        )
    except PermissionError as exc:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        return _blocked_write_result(
            _result_path(path, resolved),
            summary=str(exc).strip() or _nonwritable_target_message(resolved),
            suggested_next_step=_nonwritable_target_next_step(resolved),
            operation="write_sheet_data",
            file_type="xlsx",
            sheet=target_sheet,
            cells_written=count,
        )
    except Exception as e:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        return json.dumps({"error": str(e)}, ensure_ascii=False)
