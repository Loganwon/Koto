"""Pure workbook-structure collection helpers for XLSX task tools."""

from __future__ import annotations

import re
from typing import Any, Dict, List


def extract_external_link_targets(workbook: Any) -> List[str]:
    targets: List[str] = []
    for link in list(getattr(workbook, "_external_links", []) or []):
        relation = getattr(link, "file_link", None)
        target = (
            getattr(relation, "Target", None)
            or getattr(relation, "target", None)
            or getattr(link, "target", None)
        )
        target_text = str(target or "").strip()
        if target_text and target_text not in targets:
            targets.append(target_text)
    return targets


def _trim_trailing_empty(values: List[Any]) -> List[Any]:
    trimmed = list(values)
    while trimmed and trimmed[-1] in (None, ""):
        trimmed.pop()
    return trimmed


def _normalize_preview_value(value: Any) -> Any:
    return "" if value is None else value


def sample_sheet_rows(
    worksheet: Any,
    *,
    max_rows: int,
    max_cols: int = 12,
) -> List[Dict[str, Any]]:
    samples: List[Dict[str, Any]] = []
    limit_rows = min(max(getattr(worksheet, "max_row", 0), 0), 200)
    limit_cols = min(max(getattr(worksheet, "max_column", 0), 0), max_cols)
    if limit_rows < 1 or limit_cols < 1:
        return samples
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=limit_rows, max_col=limit_cols, values_only=True),
        start=1,
    ):
        trimmed = _trim_trailing_empty(list(row))
        if not any(value not in (None, "") for value in trimmed):
            continue
        samples.append(
            {"row": row_index, "values": [_normalize_preview_value(value) for value in trimmed]}
        )
        if len(samples) >= max_rows:
            break
    return samples


def detect_year_header(worksheet: Any, *, max_scan_rows: int = 10, max_scan_cols: int = 40) -> Dict[str, Any]:
    from openpyxl.utils import get_column_letter
    pattern = re.compile(r"(?<!\d)(20\d{2}(?:\s*[A-Za-z]{0,4})?)(?!\d)", re.IGNORECASE)
    best_row, best_matches = 0, []
    for row in worksheet.iter_rows(min_row=1, max_row=min(getattr(worksheet, "max_row", 0), max_scan_rows), max_col=min(getattr(worksheet, "max_column", 0), max_scan_cols)):
        matches = [{"index": int(cell.column), "letter": get_column_letter(int(cell.column)), "header": str(cell.value or "").strip()} for cell in row if str(cell.value or "").strip() and len(str(cell.value or "").strip()) <= 24 and pattern.search(str(cell.value or "").strip())]
        if len(matches) > len(best_matches): best_row, best_matches = int(row[0].row), matches
    return {"row": best_row, "columns": best_matches} if len(best_matches) >= 2 else {}


def collect_formula_examples(worksheet: Any, *, max_formula_examples: int) -> Dict[str, Any]:
    formulas, external = [], []
    formula_count = external_count = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if getattr(cell, "data_type", "") != "f": continue
            formula_count += 1; item = {"cell": cell.coordinate, "formula": str(cell.value or "")[:200]}
            if len(formulas) < max_formula_examples: formulas.append(item)
            if "[" in str(cell.value or ""):
                external_count += 1
                if len(external) < max_formula_examples: external.append(item)
    return {"formula_count": formula_count, "formula_examples": formulas, "external_formula_count": external_count, "external_formula_examples": external}


def row_label_for_year_series(worksheet: Any, row_index: int, *, before_column: int) -> str:
    labels = [str(worksheet.cell(row=row_index, column=index).value or "").strip() for index in range(1, max(before_column, 1))]
    return next((label for label in reversed(labels) if label), "")


def severity_for_financial_label(label: str, high_priority_hints: tuple[str, ...]) -> str:
    return "high" if any(hint in label for hint in high_priority_hints) else "medium"


def display_series_value(value: Any, formula_text: str) -> Any:
    return value if value not in (None, "") else formula_text[:160] if formula_text else ""
