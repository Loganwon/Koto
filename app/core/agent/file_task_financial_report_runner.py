from __future__ import annotations

import json
import logging
import os
import re
import tempfile
import uuid
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from app.core.agent.file_task_contract import (
    FileTaskClassification,
    FileTaskEvent,
    FileTaskFile,
    FileTaskLedger,
    FileTaskRequest,
)
from app.core.agent.file_task_recipes import (
    request_file_types,
    request_target_file_type,
    select_task_recipe,
    semantic_markers,
)

logger = logging.getLogger(__name__)


def _preview(value: Any, limit: int = 700) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 1)].rstrip() + "…"


class FileTaskFinancialReportRunner:
    """Native Excel-financial-analysis -> DOCX report workflow.

    This runner owns the deterministic multi-file financial report chain. The
    general FileTaskRuntime supplies shared host services: tool gateway,
    verification, event payload helpers, and chart/report helper functions.
    """

    recipe_id = "financial_xlsx_docx_report"

    def __init__(self, host: Any):
        self._host = host
        self._workspace_root = str(getattr(host, "_workspace_root", "") or "")

    def _display_path(self, value: Any) -> str:
        return self._host._display_path(value)

    def _resolve_task_file_path(self, path: Any) -> str:
        return self._host._resolve_task_file_path(path)

    def _prepare_matplotlib_config_dir(self) -> None:
        config_root = (
            Path(self._workspace_root or tempfile.gettempdir())
            / ".koto_artifacts"
            / "matplotlib"
        )
        config_root.mkdir(parents=True, exist_ok=True)
        os.environ["MPLCONFIGDIR"] = str(config_root)

    def _call_model(self, **kwargs: Any) -> Any:
        return self._host._call_model(**kwargs)

    def _normalize_model_response(self, response: Any, tools: List[Dict[str, Any]]):
        return self._host._normalize_model_response(response, tools)

    def _primary_financial_xlsx_file(
        self, files: List[FileTaskFile]
    ) -> Optional[FileTaskFile]:
        if not files:
            return None
        primary_markers = (
            "financial",
            "finance",
            "model",
            "p&l",
            "profit",
            "income",
            "财务",
            "预测",
            "模型",
            "损益",
            "利润",
        )
        supplemental_markers = ("销售", "台账", "流水", "sales", "ledger", "detail")

        def score(file_info: FileTaskFile, index: int) -> int:
            text = f"{file_info.name or ''} {file_info.path or ''}".lower()
            value = 0
            value += sum(4 for marker in primary_markers if marker in text)
            value -= sum(3 for marker in supplemental_markers if marker in text)
            value -= index
            return value

        ranked = sorted(
            enumerate(files),
            key=lambda item: score(item[1], item[0]),
            reverse=True,
        )
        return ranked[0][1]

    def should_route(
        self, request: FileTaskRequest, files: List[FileTaskFile]
    ) -> bool:
        recipe_match = select_task_recipe(
            request, files, write_intent=self._host._has_write_intent(request.task)
        )
        if recipe_match and recipe_match.recipe.id == self.recipe_id:
            return True
        followup_context = self._host._followup_context(request)
        previous_recipe = str(
            followup_context.get("previous_task_selected_recipe")
            or followup_context.get("previous_task_execution_mode")
            or ""
        ).strip()
        if previous_recipe != self.recipe_id:
            return False
        markers = semantic_markers(
            request.task,
            file_types=request_file_types(files),
            target_file_type=request_target_file_type(request, files),
        )
        return (
            "xlsx" in request_file_types(files)
            and markers.get("docx_target", False)
            and self._host._has_write_intent(request.task)
            and (
                markers.get("chart_request", False)
                or markers.get("problem_analysis_request", False)
                or markers.get("table_request", False)
            )
        )

    def stream(
        self,
        request: FileTaskRequest,
        context_files: List[FileTaskFile],
    ) -> Iterable[FileTaskEvent]:
        host = self._host
        ledger = FileTaskLedger(request.run_id)
        gateway = host._build_tool_gateway(request, context_files)
        executor = gateway.execute
        xlsx_files = host._context_files_by_type(context_files, {"xlsx", "xlsm"})
        xlsx_file = self._primary_financial_xlsx_file(xlsx_files)
        supplemental_xlsx_files = [
            file_info for file_info in xlsx_files if file_info is not xlsx_file
        ]
        target_docx_file = host._first_context_file(
            context_files, {"docx"}, target=True
        )
        single_docx_file = host._single_context_file(context_files, {"docx"})
        ambiguous_docx_target = (
            target_docx_file is None
            and single_docx_file is None
            and len(host._context_files_by_type(context_files, {"docx"})) > 1
        )
        docx_file = target_docx_file or single_docx_file
        xlsx_path = (
            str(xlsx_file.path or xlsx_file.name or "").strip() if xlsx_file else ""
        )
        docx_path = str(
            request.target_path
            or (docx_file.path if docx_file else "")
            or (docx_file.name if docx_file else "")
        ).strip()
        if xlsx_path:
            xlsx_path = str(Path(xlsx_path).resolve())
        if docx_path:
            docx_path = str(Path(docx_path).resolve())

        classification = FileTaskClassification(
            request_kind="new_task",
            task_family="financial_report",
            operation_kind="analyze_visualize_write",
            execution_mode=self.recipe_id,
            output_mode="write",
            write_intent=True,
            target_file_type="docx",
            file_types=sorted(host._file_types(context_files)),
            matched_capabilities=[
                "inspect_workbook_structure",
                "audit_financial_workbook",
                "run_python_code",
                "write_docx_content",
                "insert_image_into_docx",
            ],
            reason_codes=["native_financial_xlsx_docx_report", "write_intent"],
            selected_recipe=self.recipe_id,
            recipe_candidates=[
                {
                    "recipe_id": self.recipe_id,
                    "score": 1,
                    "task_family": "financial_report",
                    "operation_kind": "analyze_visualize_write",
                    "reason_codes": [f"recipe:{self.recipe_id}"],
                }
            ],
            confidence=1.0,
        )
        classification_payload = classification.public_dict()
        constraint_audit = host._financial_constraint_audit(
            request,
            context_files,
            ambiguous_docx_target=ambiguous_docx_target,
        )
        plan_runtime = host._build_runtime_metadata(
            terminal_status="plan_checked",
            readonly_fallback_used=False,
            model_failed=False,
            planner_payload={
                "backend": "native",
                "source": "native",
                "policy": "native_only",
                "transport": "internal",
                "reason": self.recipe_id,
                "round": 1,
            },
            planner_fallback_payload={},
        )

        yield ledger.event(
            "run.started",
            {
                "task": request.task,
                "mode": "whitebox_v1",
                "file_count": len(context_files),
                "target_path": docx_path,
                "model_mode": request.model_mode,
                "model_id": request.model_id,
                "constraint_audit": constraint_audit,
                **classification_payload,
            },
        )
        yield ledger.event(
            "task.classified",
            {
                **classification_payload,
                "summary": "已识别为 Excel 财务分析、图表生成与 Word 写回的原生白盒工作流。",
            },
            step_id="plan",
        )
        yield ledger.event(
            "plan.checked",
            {
                "passed": bool(xlsx_path and docx_path),
                "status": "pass" if xlsx_path and docx_path else "failed",
                "summary": (
                    "已匹配 Excel 财务分析、图表生成和 DOCX 写回工作流。"
                    if xlsx_path and docx_path
                    else (
                        "存在多个 DOCX，需明确目标 Word 文件。"
                        if ambiguous_docx_target
                        else "缺少 Excel 或 DOCX 目标文件。"
                    )
                ),
                "requirements": {"write_required": True, "target_file_type": "docx"},
                "violations": (
                    []
                    if xlsx_path and docx_path
                    else (
                        ["ambiguous_docx_target"]
                        if ambiguous_docx_target
                        else ["missing_xlsx_or_docx_target"]
                    )
                ),
                "runtime": plan_runtime,
                "constraint_audit": constraint_audit,
            },
            step_id="plan",
        )

        if not xlsx_path or not docx_path:
            terminal_runtime = host._build_runtime_metadata(
                terminal_status="failed",
                readonly_fallback_used=False,
                model_failed=False,
                planner_payload=plan_runtime.get("planner", {}),
                planner_fallback_payload={},
            )
            summary = (
                "存在多个 DOCX，需明确目标 Word 文件，无法自动写入财务图表。"
                if ambiguous_docx_target
                else "缺少 Excel 或 DOCX 目标文件，无法生成并写入财务图表。"
            )
            yield ledger.event(
                "run.finished",
                {
                    "task": request.task,
                    "mode": "whitebox_v1",
                    "summary": summary,
                    "completed_task": False,
                    "context": [],
                    "file_changes": [],
                    "runtime": terminal_runtime,
                    "constraint_audit": constraint_audit,
                    **classification_payload,
                },
            )
            return

        yield ledger.event(
            "plan.created",
            {
                "summary": f"准备分析 {host._display_path(xlsx_path)}{('，并联动 ' + str(len(supplemental_xlsx_files)) + ' 份补充 Excel') if supplemental_xlsx_files else ''}，生成图表和问题清单，并写入 {host._display_path(docx_path)}。",
                "steps": [
                    {
                        "id": "context",
                        "title": "读取财务模型",
                        "description": "检查主财务模型结构、外部链接、公式和关键工作表；如有补充 Excel，同步读取其表结构。",
                    },
                    {
                        "id": "execute",
                        "title": "生成图表和问题清单",
                        "description": "抽取关键年份指标和补充销售台账指标，生成 PNG 图表，并结合审计结果整理问题清单。",
                    },
                    {
                        "id": "write_docx",
                        "title": "写入 Word",
                        "description": "先写入问题清单，再插入真实图表图片。",
                    },
                    {
                        "id": "check",
                        "title": "核验结果",
                        "description": "确认目标 DOCX 已产生文件变更。",
                    },
                ],
                "success_criteria": [
                    "目标 DOCX 产生 file.changed 事件",
                    "图表作为真实图片插入 DOCX",
                    "问题清单作为可读段落写入 DOCX",
                ],
                "constraint_audit": constraint_audit,
            },
        )

        file_changes: List[Dict[str, Any]] = []
        snippets: List[Dict[str, Any]] = []
        model_failed = False
        readonly_fallback_used = False
        tool_runtime_outcome: Optional[Dict[str, Any]] = None

        yield ledger.event(
            "step.started",
            {
                "title": "读取财务模型",
                "detail": "使用专用 Excel 审计工具读取结构和问题线索。",
            },
            step_id="context",
        )
        inspect_payload, inspect_events = host._run_builtin_tool(
            ledger,
            executor,
            step_id="context",
            tool_name="inspect_workbook_structure",
            tool_args={
                "path": xlsx_path,
                "sample_rows_per_sheet": 8,
                "max_formula_examples_per_sheet": 8,
            },
            file_changes=file_changes,
        )
        yield from inspect_events
        audit_payload, audit_events = host._run_builtin_tool(
            ledger,
            executor,
            step_id="context",
            tool_name="audit_financial_workbook",
            tool_args={
                "path": xlsx_path,
                "sample_rows_per_sheet": 6,
                "max_formula_examples_per_sheet": 8,
                "max_findings": 12,
            },
            file_changes=file_changes,
        )
        yield from audit_events
        snippets.append(
            {
                "source": host._display_path(xlsx_path),
                "path": xlsx_path,
                "preview": _preview(
                    (audit_payload or {}).get("summary")
                    or (inspect_payload or {}).get("summary")
                    or "已读取财务模型",
                    500,
                ),
                "chars": 0,
            }
        )
        supplemental_inspections: List[Dict[str, Any]] = []
        for supplemental_file in supplemental_xlsx_files:
            supplemental_path = str(
                supplemental_file.path or supplemental_file.name or ""
            ).strip()
            if not supplemental_path:
                continue
            supplemental_path = str(Path(supplemental_path).resolve())
            supplemental_payload, supplemental_events = host._run_builtin_tool(
                ledger,
                executor,
                step_id="context",
                tool_name="inspect_workbook_structure",
                tool_args={
                    "path": supplemental_path,
                    "sample_rows_per_sheet": 8,
                    "max_formula_examples_per_sheet": 8,
                },
                file_changes=file_changes,
            )
            yield from supplemental_events
            supplemental_inspections.append(
                {
                    "file": supplemental_file,
                    "path": supplemental_path,
                    "inspect_payload": supplemental_payload,
                }
            )
            snippets.append(
                {
                    "source": host._display_path(supplemental_path),
                    "path": supplemental_path,
                    "preview": _preview(
                        (supplemental_payload or {}).get("summary")
                        or "已读取补充 Excel",
                        500,
                    ),
                    "chars": 0,
                }
            )
        yield ledger.event(
            "step.finished", {"summary": "已完成财务模型结构检查。"}, step_id="context"
        )
        yield ledger.event(
            "step.result",
            host._build_step_result_payload(
                title="读取财务模型",
                summary="已完成财务模型结构检查，并收集审计问题线索。",
                status="completed",
                snippet_count=len(snippets),
                snippets=snippets,
            ),
            step_id="context",
        )

        yield ledger.event(
            "step.started",
            {
                "title": "生成图表和问题清单",
                "detail": "从 Excel 抽取关键指标，生成可插入 Word 的 PNG 图表。",
            },
            step_id="execute",
        )
        chart_result = self._generate_financial_workbook_chart(
            xlsx_path, inspect_payload, audit_payload
        )
        yield from self._emit_chart_pipeline_events(
            ledger,
            step_id="execute",
            code="internal_financial_workbook_chart_pipeline",
            result=chart_result,
            default_message="已生成财务图表。",
        )
        supplemental_chart_results: List[Dict[str, Any]] = []
        for supplemental in supplemental_inspections:
            supplemental_path = str(supplemental.get("path") or "").strip()
            if not supplemental_path:
                continue
            supplemental_result = self._generate_supplemental_xlsx_chart(
                supplemental_path,
                supplemental.get("inspect_payload")
                if isinstance(supplemental.get("inspect_payload"), dict)
                else None,
                display_name=str(
                    getattr(supplemental.get("file"), "name", "") or ""
                ).strip(),
            )
            supplemental_chart_results.append(supplemental_result)
            yield from self._emit_chart_pipeline_events(
                ledger,
                step_id="execute",
                code="internal_supplemental_xlsx_chart_pipeline",
                result=supplemental_result,
                default_message="补充 Excel 图表处理完成。",
                source_path=supplemental_path,
            )

        problems = self._financial_report_problem_paragraphs(
            audit_payload, inspect_payload, chart_result
        )
        for supplemental_result in supplemental_chart_results:
            supplemental_paragraphs = supplemental_result.get("paragraphs")
            if isinstance(supplemental_paragraphs, list):
                problems.extend(
                    item
                    for item in supplemental_paragraphs
                    if isinstance(item, dict) and str(item.get("text") or "").strip()
                )
            elif not supplemental_result.get("success"):
                problems.extend(
                    [
                        {"text": "补充 Excel 分析", "style": "Heading 2"},
                        {
                            "text": f"补充 Excel 图表生成未成功：{supplemental_result.get('error') or '缺少可作图数据'}。",
                            "style": "List Bullet",
                        },
                    ]
                )
        model_synthesis = self._financial_report_model_synthesis(
            request, audit_payload, inspect_payload, chart_result
        )
        if model_synthesis:
            yield ledger.event(
                "tool.finished",
                {
                    "tool_name": "model_message",
                    "success": True,
                    "result_preview": _preview(model_synthesis, 900),
                },
                step_id="execute",
            )
            problems = self._merge_financial_model_synthesis(
                problems, model_synthesis
            )
        yield ledger.event(
            "step.finished",
            {"summary": "已生成图表并整理问题清单。"},
            step_id="execute",
        )
        yield ledger.event(
            "step.result",
            host._build_step_result_payload(
                title="生成图表和问题清单",
                summary=(
                    "已生成图表并整理问题清单。"
                    if chart_result.get("success")
                    else "已整理问题清单，但图表生成失败。"
                ),
                status=(
                    "completed" if chart_result.get("success") else "needs_attention"
                ),
                file_changes=file_changes,
            ),
            step_id="execute",
        )

        yield ledger.event(
            "step.started",
            {
                "title": "写入 Word",
                "detail": "把问题清单和图表写入目标 DOCX。",
            },
            step_id="write_docx",
        )
        _, write_events = host._run_builtin_tool(
            ledger,
            executor,
            step_id="write_docx",
            tool_name="write_docx_content",
            tool_args={
                "path": docx_path,
                "paragraphs": json.dumps(problems, ensure_ascii=False),
            },
            file_changes=file_changes,
        )
        yield from write_events
        if chart_result.get("success") and chart_result.get("path"):
            _, image_events = host._run_builtin_tool(
                ledger,
                executor,
                step_id="write_docx",
                tool_name="insert_image_into_docx",
                tool_args={
                    "path": docx_path,
                    "image_path": str(chart_result.get("path") or ""),
                    "title": "关键财务指标趋势图",
                    "caption": chart_result.get("caption")
                    or "根据 Excel 财务模型关键年份数据自动生成。",
                    "width_inches": 6.5,
                },
                file_changes=file_changes,
            )
            yield from image_events
        for index, supplemental_result in enumerate(supplemental_chart_results, start=1):
            if not supplemental_result.get("success") or not supplemental_result.get(
                "path"
            ):
                continue
            _, image_events = host._run_builtin_tool(
                ledger,
                executor,
                step_id="write_docx",
                tool_name="insert_image_into_docx",
                tool_args={
                    "path": docx_path,
                    "image_path": str(supplemental_result.get("path") or ""),
                    "title": (
                        f"补充 Excel 数据图 {index}"
                        if len(supplemental_chart_results) > 1
                        else "补充 Excel 数据图"
                    ),
                    "caption": supplemental_result.get("caption")
                    or "根据补充 Excel 台账数据自动生成。",
                    "width_inches": 6.5,
                },
                file_changes=file_changes,
            )
            yield from image_events
        yield ledger.event(
            "step.finished",
            {
                "title": "写入 Word 完成",
                "summary": f"已记录 {len(file_changes)} 次文件变更。",
            },
            step_id="write_docx",
        )
        yield ledger.event(
            "step.result",
            host._build_step_result_payload(
                title="写入 Word",
                summary=("已将问题清单和图表写入 Word。" if file_changes else "未检测到 Word 文件变更。"),
                status="completed" if file_changes else "failed",
                file_changes=file_changes,
            ),
            step_id="write_docx",
        )

        yield ledger.event(
            "check.started",
            {
                "title": "检查执行状态",
                "criteria": host._success_criteria(request, True, "write"),
            },
            step_id="check",
        )
        check_payload = host._verify_task(
            request,
            executor,
            file_changes,
            True,
            "write",
            model_failed,
            readonly_fallback_used,
            tool_runtime_outcome,
            None,
            None,
        )
        terminal_runtime = host._build_runtime_metadata(
            terminal_status=str(check_payload.get("status") or "").strip(),
            readonly_fallback_used=readonly_fallback_used,
            model_failed=model_failed,
            planner_payload={
                "backend": "native",
                "source": "native",
                "policy": "native_only",
                "transport": "internal",
                "reason": self.recipe_id,
                "round": 1,
            },
            planner_fallback_payload={},
        )
        check_payload["runtime"] = terminal_runtime
        yield ledger.event("check.finished", check_payload, step_id="check")
        yield ledger.event(
            "step.result",
            host._build_step_result_payload(
                title="检查执行状态",
                summary=str(check_payload.get("summary") or "检查完成。"),
                status=host._check_step_result_status(check_payload),
                runtime=terminal_runtime,
                passed=check_payload.get("passed"),
                file_changes=file_changes,
            ),
            step_id="check",
        )
        summary = (
            (
                f"已分析 Excel 财务模型，并联动 {len(supplemental_chart_results)} 份补充 Excel，生成图表并把问题清单写入 Word。"
                if supplemental_chart_results
                else "已分析 Excel 财务模型，生成图表并把问题清单写入 Word。"
            )
            if check_payload.get("passed")
            else str(check_payload.get("summary") or "未能确认目标 DOCX 已更新。")
        )
        yield ledger.event(
            "run.finished",
            {
                "task": request.task,
                "mode": "whitebox_v1",
                "summary": summary,
                "completed_task": bool(check_payload.get("passed")),
                "context": snippets[:8],
                "file_changes": file_changes,
                "runtime": terminal_runtime,
                "constraint_audit": constraint_audit,
                **classification_payload,
            },
        )

    def _financial_cell_number(self, value: Any) -> Optional[float]:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            number = float(value)
            return number if number == number else None
        text = str(value or "").strip().replace(",", "")
        if not text:
            return None
        is_percent = text.endswith("%")
        text = text.rstrip("%")
        try:
            number = float(text)
        except Exception:
            return None
        if is_percent:
            number = number / 100.0
        return number if number == number else None

    def _financial_year_label(self, value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        match = re.search(r"(20\d{2}\s*[AE]?)", text, re.IGNORECASE)
        return match.group(1).replace(" ", "").upper() if match else ""

    def _find_financial_sheet_rows(
        self, workbook: Any
    ) -> tuple[str, List[tuple[Any, ...]], List[tuple[int, str]]]:
        preferred = [
            name
            for name in workbook.sheetnames
            if re.search(r"(?:p&l|profit|income|利润|损益)", name, re.IGNORECASE)
        ]
        candidates = preferred or list(workbook.sheetnames)
        for sheet_name in candidates:
            worksheet = workbook[sheet_name]
            rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
            for row in rows[:20]:
                year_columns = [
                    (idx, label)
                    for idx, cell in enumerate(row)
                    if (label := self._financial_year_label(cell))
                ]
                if len(year_columns) >= 2:
                    return sheet_name, rows, year_columns
        first = workbook.sheetnames[0] if workbook.sheetnames else ""
        return (
            first,
            (
                [tuple(row) for row in workbook[first].iter_rows(values_only=True)]
                if first
                else []
            ),
            [],
        )

    def _financial_row_label(
        self, row: tuple[Any, ...], year_columns: List[tuple[int, str]]
    ) -> str:
        first_year_index = min([idx for idx, _ in year_columns] or [3])
        labels = [
            str(cell or "").strip()
            for cell in row[:first_year_index]
            if str(cell or "").strip()
        ]
        return labels[-1] if labels else ""

    def _financial_row_values(
        self, row: tuple[Any, ...], year_columns: List[tuple[int, str]]
    ) -> List[Optional[float]]:
        return [
            self._financial_cell_number(row[idx] if idx < len(row) else None)
            for idx, _ in year_columns
        ]

    def _extract_financial_series_groups(
        self,
        rows: List[tuple[Any, ...]],
        year_columns: List[tuple[int, str]],
    ) -> Dict[str, Dict[str, List[Optional[float]]]]:
        groups: Dict[str, Dict[str, List[Optional[float]]]] = {
            "money": {},
            "rates": {},
            "volume": {},
            "expenses": {},
            "product_revenue": {},
            "costs": {},
        }
        exact_map = {
            "收入合计": ("money", "收入合计"),
            "硬件收入": ("money", "硬件收入"),
            "配件收入": ("money", "配件收入"),
            "互联网业务收入": ("money", "互联网业务收入"),
            "成本合计": ("costs", "成本合计"),
            "硬件成本": ("costs", "硬件成本"),
            "配件成本": ("costs", "配件成本"),
            "互联网业务成本": ("costs", "互联网业务成本"),
            "毛利合计": ("money", "毛利合计"),
            "费用合计": ("expenses", "费用合计"),
            "研发费用": ("expenses", "研发费用"),
            "销售费用": ("expenses", "销售费用"),
            "管理费用": ("expenses", "管理费用"),
            "财务费用": ("expenses", "财务费用"),
            "利润总额": ("money", "利润总额"),
            "净利润": ("money", "净利润"),
            "增速%": ("rates", "收入增速"),
            "综合毛利率%": ("rates", "综合毛利率"),
            "硬件整体毛利率%": ("rates", "硬件毛利率"),
            "净利率%": ("rates", "净利率"),
            "研发费用%": ("rates", "研发费用率"),
            "销售费用%": ("rates", "销售费用率"),
            "管理费用%": ("rates", "管理费用率"),
            "销量": ("volume", "总销量"),
        }
        section = ""
        for row in rows:
            label = self._financial_row_label(row, year_columns)
            if not label:
                continue
            values = self._financial_row_values(row, year_columns)
            if sum(value is not None for value in values) < 2:
                if label in {"销量", "销量%", "硬件收入", "硬件成本"}:
                    section = label
                continue
            if label in {"销量", "销量%", "硬件收入", "硬件成本"}:
                section = label
            mapped = exact_map.get(label)
            if mapped:
                group, display = mapped
                groups[group].setdefault(display, values)
                continue
            if label in {"XR系列", "AI系列", "AR系列"}:
                product = label.replace("系列", "")
                if section == "销量":
                    groups["volume"].setdefault(f"{product}销量", values)
                elif section == "硬件收入":
                    groups["product_revenue"].setdefault(f"{product}收入", values)
                elif section == "硬件成本":
                    groups["costs"].setdefault(f"{product}成本", values)
        if not any(groups.values()):
            fallback: Dict[str, List[Optional[float]]] = {}
            for row in rows:
                label = self._financial_row_label(row, year_columns)
                values = self._financial_row_values(row, year_columns)
                if label and sum(value is not None for value in values) >= 2:
                    fallback[label[:20]] = values
                if len(fallback) >= 4:
                    break
            groups["money"] = fallback
        return {key: value for key, value in groups.items() if value}

    def _flatten_financial_series(
        self, groups: Dict[str, Dict[str, List[Optional[float]]]]
    ) -> Dict[str, List[Optional[float]]]:
        flat: Dict[str, List[Optional[float]]] = {}
        for group in (
            "money",
            "rates",
            "volume",
            "expenses",
            "product_revenue",
            "costs",
        ):
            for name, values in (groups.get(group) or {}).items():
                flat.setdefault(name, values)
        return flat

    def _generate_financial_workbook_chart(
        self,
        xlsx_path: str,
        inspect_payload: Optional[Dict[str, Any]],
        audit_payload: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        resolved = self._resolve_task_file_path(xlsx_path)
        if not resolved:
            return {"success": False, "error": f"无法定位 Excel 文件：{xlsx_path}"}
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
            sheet_name, rows, year_columns = self._find_financial_sheet_rows(workbook)
            workbook.close()
            if not year_columns:
                return {"success": False, "error": "未识别到可用于作图的年份列。"}
            years = [label for _, label in year_columns]
            series_groups = self._extract_financial_series_groups(rows, year_columns)
            series = self._flatten_financial_series(series_groups)
            if not series:
                return {
                    "success": False,
                    "error": "未识别到可用于作图的关键财务指标行。",
                }
            try:
                self._prepare_matplotlib_config_dir()
                import matplotlib

                matplotlib.use("Agg")
                import matplotlib.pyplot as plt
                from matplotlib import font_manager
            except ModuleNotFoundError as exc:
                if exc.name != "matplotlib":
                    raise
                return self._generate_financial_workbook_chart_pil(
                    xlsx_path=xlsx_path,
                    sheet_name=sheet_name,
                    years=[label for _, label in year_columns],
                    series_groups=series_groups,
                    series=series,
                )

            available_fonts = {font.name for font in font_manager.fontManager.ttflist}
            for font_name in (
                "Microsoft YaHei",
                "SimHei",
                "Noto Sans CJK SC",
                "WenQuanYi Micro Hei",
                "DejaVu Sans",
            ):
                if font_name in available_fonts:
                    plt.rcParams["font.sans-serif"] = [font_name]
                    break
            plt.rcParams["axes.unicode_minus"] = False

            fig, axes = plt.subplots(2, 2, figsize=(12.5, 8.4))
            fig.suptitle(f"{sheet_name} 财务预测质量检查", fontsize=15, fontweight="bold")

            def plot_lines(
                ax: Any,
                data: Dict[str, List[Optional[float]]],
                title: str,
                ylabel: str,
                *,
                as_percent: bool = False,
            ) -> int:
                plotted_count = 0
                for name, values in data.items():
                    numeric_values = [
                        float(value) if value is not None else None for value in values
                    ]
                    if sum(value is not None for value in numeric_values) < 2:
                        continue
                    y_values = [
                        value * 100 if (as_percent and value is not None) else value
                        for value in numeric_values
                    ]
                    ax.plot(years, y_values, marker="o", linewidth=2, label=name)
                    plotted_count += 1
                ax.set_title(title)
                ax.set_xlabel("年份")
                ax.set_ylabel(ylabel)
                ax.grid(True, alpha=0.25)
                if plotted_count:
                    ax.legend(loc="best", fontsize=8)
                else:
                    ax.text(
                        0.5,
                        0.5,
                        "未识别到足够数据",
                        ha="center",
                        va="center",
                        transform=ax.transAxes,
                        color="#666",
                    )
                return plotted_count

            monetary = {
                key: value
                for key, value in (series_groups.get("money") or {}).items()
                if key in {"收入合计", "毛利合计", "净利润", "利润总额"}
            } or (series_groups.get("money") or {})
            expenses = series_groups.get("expenses") or {}
            rates = series_groups.get("rates") or {}
            volume = series_groups.get("volume") or {}

            plotted = 0
            plotted += plot_lines(axes[0][0], monetary, "收入、毛利与利润", "人民币万元")
            plotted += plot_lines(axes[0][1], rates, "增长率与利润率", "百分比", as_percent=True)
            plotted += plot_lines(axes[1][0], volume, "销量与产品结构", "台/套")
            plotted += plot_lines(axes[1][1], expenses, "费用结构", "人民币万元")
            if not plotted:
                plt.close(fig)
                return {
                    "success": False,
                    "error": "关键财务指标有效数值不足，无法生成趋势图。",
                }
            fig.tight_layout(rect=(0, 0, 1, 0.96))
            artifact_root = (
                Path(self._workspace_root or tempfile.gettempdir())
                / ".koto_artifacts"
                / "financial_charts"
            )
            artifact_root.mkdir(parents=True, exist_ok=True)
            chart_path = artifact_root / f"financial_chart_{uuid.uuid4().hex[:10]}.png"
            fig.savefig(chart_path, dpi=240, bbox_inches="tight")
            plt.close(fig)
            chart_issues = self._financial_series_issues(series, years)
            return {
                "success": True,
                "path": str(chart_path),
                "sheet": sheet_name,
                "years": years,
                "series": series,
                "series_groups": series_groups,
                "issues": chart_issues,
                "summary": f"已从“{sheet_name}”生成 {plotted} 条指标线、4 个分析面板的财务图表：{chart_path.name}",
                "caption": f"数据来源：{self._display_path(xlsx_path)} / {sheet_name}；年份：{', '.join(years)}；图表按金额、比率、销量、费用分面展示。",
            }
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    def _pil_chart_font(self, size: int, *, bold: bool = False) -> Any:
        try:
            from PIL import ImageFont

            candidates = [
                "C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc",
                "C:/Windows/Fonts/simhei.ttf",
                "C:/Windows/Fonts/arial.ttf",
            ]
            for font_path in candidates:
                if font_path and os.path.exists(font_path):
                    return ImageFont.truetype(font_path, size=size)
            return ImageFont.load_default()
        except Exception:
            return None

    def _pil_draw_text(
        self,
        draw: Any,
        xy: tuple[int, int],
        text: Any,
        *,
        fill: str = "#111111",
        font: Any = None,
    ) -> None:
        clean = str(text or "").strip()
        try:
            draw.text(xy, clean, fill=fill, font=font)
        except Exception:
            draw.text(
                xy,
                clean.encode("ascii", "ignore").decode("ascii"),
                fill=fill,
                font=font,
            )

    def _save_pil_chart_image(
        self,
        chart_path: Path,
        *,
        title: str,
        panels: List[Dict[str, Any]],
    ) -> None:
        from PIL import Image, ImageDraw

        width, height = 1600, 980
        image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(image)
        title_font = self._pil_chart_font(30, bold=True)
        panel_font = self._pil_chart_font(22, bold=True)
        text_font = self._pil_chart_font(18)
        small_font = self._pil_chart_font(15)
        self._pil_draw_text(draw, (48, 28), title, font=title_font)
        colors = ["#4C78A8", "#F58518", "#54A24B", "#E45756", "#72B7B2", "#B279A2"]
        panel_boxes = [
            (55, 105, 770, 475),
            (830, 105, 1545, 475),
            (55, 545, 770, 915),
            (830, 545, 1545, 915),
        ]

        def numeric_values(values: Any) -> List[float]:
            parsed: List[float] = []
            for value in values or []:
                number = self._financial_cell_number(value)
                if number is not None:
                    parsed.append(float(number))
            return parsed

        for panel_index, box in enumerate(panel_boxes):
            if panel_index >= len(panels):
                break
            panel = panels[panel_index]
            left, top, right, bottom = box
            draw.rectangle(box, outline="#D9DDE7", width=2)
            self._pil_draw_text(
                draw,
                (left + 18, top + 14),
                panel.get("title") or "Chart",
                font=panel_font,
            )
            plot_left, plot_top = left + 70, top + 65
            plot_right, plot_bottom = right - 32, bottom - 62
            draw.line((plot_left, plot_bottom, plot_right, plot_bottom), fill="#777")
            draw.line((plot_left, plot_top, plot_left, plot_bottom), fill="#777")
            kind = panel.get("kind") or "line"
            if kind == "bar":
                bars = [
                    (str(label), float(value))
                    for label, value in panel.get("bars") or []
                    if self._financial_cell_number(value) is not None
                ][:10]
                if not bars:
                    self._pil_draw_text(
                        draw,
                        (plot_left + 120, plot_top + 95),
                        "No plottable data",
                        fill="#666666",
                        font=text_font,
                    )
                    continue
                max_value = max(abs(value) for _, value in bars) or 1.0
                row_height = max(18, int((plot_bottom - plot_top) / max(len(bars), 1)))
                for idx, (label, value) in enumerate(bars):
                    y = plot_top + idx * row_height + 6
                    bar_width = int((plot_right - plot_left - 165) * abs(value) / max_value)
                    draw.rectangle(
                        (plot_left + 150, y, plot_left + 150 + bar_width, y + row_height - 8),
                        fill=colors[idx % len(colors)],
                    )
                    self._pil_draw_text(
                        draw,
                        (plot_left, y),
                        label[:14],
                        font=small_font,
                    )
                    self._pil_draw_text(
                        draw,
                        (plot_left + 156 + bar_width, y),
                        f"{value:,.0f}",
                        font=small_font,
                    )
                continue

            labels = [str(label) for label in panel.get("labels") or []]
            series = {
                str(name): numeric_values(values)
                for name, values in (panel.get("series") or {}).items()
            }
            series = {name: values for name, values in series.items() if len(values) >= 2}
            if not series or not labels:
                self._pil_draw_text(
                    draw,
                    (plot_left + 120, plot_top + 95),
                    "No plottable data",
                    fill="#666666",
                    font=text_font,
                )
                continue
            all_values = [value for values in series.values() for value in values]
            min_value = min(all_values)
            max_value = max(all_values)
            if min_value == max_value:
                min_value -= 1
                max_value += 1
            point_count = max(len(labels), max(len(values) for values in series.values()))
            x_span = max(1, point_count - 1)
            for idx, label in enumerate(labels[:point_count]):
                x = int(plot_left + (plot_right - plot_left) * idx / x_span)
                self._pil_draw_text(
                    draw,
                    (x - 22, plot_bottom + 10),
                    label[:8],
                    font=small_font,
                )
            for series_index, (name, values) in enumerate(series.items()):
                points: List[tuple[int, int]] = []
                for idx, value in enumerate(values[:point_count]):
                    x = int(plot_left + (plot_right - plot_left) * idx / x_span)
                    y = int(
                        plot_bottom
                        - (plot_bottom - plot_top)
                        * (value - min_value)
                        / (max_value - min_value)
                    )
                    points.append((x, y))
                color = colors[series_index % len(colors)]
                if len(points) >= 2:
                    draw.line(points, fill=color, width=4)
                for point in points:
                    draw.ellipse(
                        (point[0] - 5, point[1] - 5, point[0] + 5, point[1] + 5),
                        fill=color,
                    )
                legend_y = plot_top + series_index * 22
                draw.rectangle(
                    (plot_right - 150, legend_y, plot_right - 132, legend_y + 12),
                    fill=color,
                )
                self._pil_draw_text(
                    draw,
                    (plot_right - 126, legend_y - 5),
                    name[:14],
                    font=small_font,
                )
        chart_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(chart_path)

    def _generate_financial_workbook_chart_pil(
        self,
        *,
        xlsx_path: str,
        sheet_name: str,
        years: List[str],
        series_groups: Dict[str, Dict[str, List[Optional[float]]]],
        series: Dict[str, List[Optional[float]]],
    ) -> Dict[str, Any]:
        artifact_root = (
            Path(self._workspace_root or tempfile.gettempdir())
            / ".koto_artifacts"
            / "financial_charts"
        )
        chart_path = artifact_root / f"financial_chart_{uuid.uuid4().hex[:10]}.png"
        monetary = {
            key: value
            for key, value in (series_groups.get("money") or {}).items()
            if key in {"收入合计", "毛利合计", "净利润", "利润总额", "Revenue", "Gross Profit", "Net Profit"}
        } or (series_groups.get("money") or {})
        panels = [
            {"title": "Revenue / Profit", "labels": years, "series": monetary},
            {"title": "Rates", "labels": years, "series": series_groups.get("rates") or {}},
            {"title": "Volume", "labels": years, "series": series_groups.get("volume") or {}},
            {"title": "Expenses", "labels": years, "series": series_groups.get("expenses") or {}},
        ]
        self._save_pil_chart_image(
            chart_path,
            title=f"{sheet_name} Financial Forecast Check",
            panels=panels,
        )
        chart_issues = self._financial_series_issues(series, years)
        return {
            "success": True,
            "path": str(chart_path),
            "sheet": sheet_name,
            "years": years,
            "series": series,
            "series_groups": series_groups,
            "issues": chart_issues,
            "summary": f"已使用 Pillow 兜底生成财务图表：{chart_path.name}",
            "caption": f"数据来源：{self._display_path(xlsx_path)} / {sheet_name}；年份：{', '.join(years)}；图表按金额、比率、销量、费用分面展示。",
        }

    def _supplemental_period_label(self, value: Any) -> str:
        if hasattr(value, "year") and hasattr(value, "month"):
            try:
                return f"{int(value.year):04d}-{int(value.month):02d}"
            except Exception:
                pass
        text = str(value or "").strip()
        if not text:
            return ""
        match = re.search(r"(20\d{2})[-/.年\s]*(\d{1,2})", text)
        if match:
            return f"{match.group(1)}-{int(match.group(2)):02d}"
        match = re.search(r"(\d{1,2})\s*月", text)
        if match:
            return f"{int(match.group(1)):02d}月"
        return text[:20]

    def _generate_supplemental_xlsx_chart(
        self,
        xlsx_path: str,
        inspect_payload: Optional[Dict[str, Any]],
        *,
        display_name: str = "",
    ) -> Dict[str, Any]:
        resolved = self._resolve_task_file_path(xlsx_path)
        if not resolved:
            return {"success": False, "error": f"无法定位 Excel 文件：{xlsx_path}"}
        display_label = str(display_name or self._display_path(xlsx_path)).strip()
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
            sheet_rows: List[tuple[str, List[tuple[Any, ...]]]] = []
            for worksheet in workbook.worksheets:
                rows = [
                    tuple(row)
                    for row in worksheet.iter_rows(max_row=500, values_only=True)
                ]
                nonempty_count = sum(
                    1 for row in rows if any(str(cell or "").strip() for cell in row)
                )
                if nonempty_count:
                    sheet_rows.append((worksheet.title, rows))
            workbook.close()
            if not sheet_rows:
                return {"success": False, "error": "补充 Excel 未识别到可读取数据行。"}
            sheet_name, rows = max(
                sheet_rows,
                key=lambda item: sum(
                    1 for row in item[1] if any(str(cell or "").strip() for cell in row)
                ),
            )

            header_index = -1
            header_score = -1
            for idx, row in enumerate(rows[:25]):
                cleaned = [str(cell or "").strip() for cell in row]
                nonempty = sum(bool(item) for item in cleaned)
                if nonempty < 2:
                    continue
                next_rows = rows[idx + 1 : idx + 8]
                numeric_cells = 0
                for next_row in next_rows:
                    numeric_cells += sum(
                        self._financial_cell_number(cell) is not None
                        for cell in next_row
                    )
                score = nonempty + numeric_cells
                if score > header_score:
                    header_score = score
                    header_index = idx
            if header_index < 0:
                return {"success": False, "error": "补充 Excel 未识别到表头行。"}

            raw_headers = list(rows[header_index])
            headers: List[str] = []
            for idx, value in enumerate(raw_headers):
                header = str(value or "").strip() or f"列{idx + 1}"
                if header in headers:
                    header = f"{header}_{idx + 1}"
                headers.append(header)
            records = [
                tuple(row)
                for row in rows[header_index + 1 :]
                if any(str(cell or "").strip() for cell in row)
            ]
            if not records:
                return {"success": False, "error": "补充 Excel 表头下方没有数据行。"}

            numeric_columns: List[tuple[int, str, int]] = []
            text_columns: List[tuple[int, str, int]] = []
            for idx, header in enumerate(headers):
                numeric_count = 0
                text_count = 0
                for row in records:
                    cell = row[idx] if idx < len(row) else None
                    if self._financial_cell_number(cell) is not None:
                        numeric_count += 1
                    elif str(cell or "").strip():
                        text_count += 1
                if numeric_count:
                    numeric_columns.append((idx, header, numeric_count))
                if text_count:
                    text_columns.append((idx, header, text_count))
            if not numeric_columns:
                return {"success": False, "error": "补充 Excel 未识别到可作图的数值列。"}

            def preferred_column(
                columns: List[tuple[int, str, int]], markers: tuple[str, ...]
            ) -> Optional[tuple[int, str, int]]:
                for column in columns:
                    lowered = column[1].lower()
                    if any(marker in lowered for marker in markers):
                        return column
                return max(columns, key=lambda item: item[2]) if columns else None

            value_column = preferred_column(
                numeric_columns,
                ("销售额", "金额", "收入", "营收", "回款", "sales", "revenue", "amount", "total"),
            ) or numeric_columns[0]
            category_column = preferred_column(
                text_columns,
                ("产品", "品类", "客户", "渠道", "地区", "区域", "product", "category", "customer", "channel", "region"),
            )
            period_column = preferred_column(
                text_columns,
                ("日期", "月份", "月", "date", "month", "period"),
            )

            value_idx, value_name, _ = value_column
            category_totals: Dict[str, float] = {}
            period_totals: Dict[str, float] = {}
            numeric_values: List[float] = []
            for row_index, row in enumerate(records, start=1):
                value = self._financial_cell_number(
                    row[value_idx] if value_idx < len(row) else None
                )
                if value is None:
                    continue
                numeric_values.append(float(value))
                if category_column:
                    category_idx = category_column[0]
                    category = str(
                        row[category_idx] if category_idx < len(row) else ""
                    ).strip()
                else:
                    category = f"记录{row_index}"
                category = category or "未命名"
                category_totals[category] = category_totals.get(category, 0.0) + float(
                    value
                )
                if period_column:
                    period_idx = period_column[0]
                    period = self._supplemental_period_label(
                        row[period_idx] if period_idx < len(row) else None
                    )
                    if period:
                        period_totals[period] = period_totals.get(period, 0.0) + float(
                            value
                        )
            if not numeric_values:
                return {"success": False, "error": "补充 Excel 数值列没有有效数值。"}

            top_categories = sorted(
                category_totals.items(), key=lambda item: item[1], reverse=True
            )[:10]
            sorted_periods = sorted(period_totals.items(), key=lambda item: item[0])
            try:
                self._prepare_matplotlib_config_dir()
                import matplotlib

                matplotlib.use("Agg")
                import matplotlib.pyplot as plt
                from matplotlib import font_manager
            except ModuleNotFoundError as exc:
                if exc.name != "matplotlib":
                    raise
                return self._generate_supplemental_xlsx_chart_pil(
                    xlsx_path=xlsx_path,
                    sheet_name=sheet_name,
                    value_name=value_name,
                    category_name=category_column[1] if category_column else "",
                    period_name=period_column[1] if period_column else "",
                    record_count=len(records),
                    numeric_values=numeric_values,
                    top_categories=top_categories,
                    sorted_periods=sorted_periods,
                    display_name=display_label,
                )

            available_fonts = {font.name for font in font_manager.fontManager.ttflist}
            for font_name in (
                "Microsoft YaHei",
                "SimHei",
                "Noto Sans CJK SC",
                "WenQuanYi Micro Hei",
                "DejaVu Sans",
            ):
                if font_name in available_fonts:
                    plt.rcParams["font.sans-serif"] = [font_name]
                    break
            plt.rcParams["axes.unicode_minus"] = False

            fig, axes = plt.subplots(1, 2, figsize=(12.5, 5.2))
            fig.suptitle(
                f"{display_label} 补充数据分析",
                fontsize=14,
                fontweight="bold",
            )
            if top_categories:
                labels = [item[0][:14] for item in top_categories]
                values = [item[1] for item in top_categories]
                axes[0].barh(labels[::-1], values[::-1], color="#4C78A8")
                axes[0].set_title(f"按{category_column[1] if category_column else '记录'}汇总")
                axes[0].set_xlabel(value_name)
                axes[0].grid(True, axis="x", alpha=0.25)
            else:
                axes[0].text(
                    0.5,
                    0.5,
                    "未识别到分类列",
                    ha="center",
                    va="center",
                    transform=axes[0].transAxes,
                    color="#666",
                )
            if len(sorted_periods) >= 2:
                axes[1].plot(
                    [item[0] for item in sorted_periods],
                    [item[1] for item in sorted_periods],
                    marker="o",
                    linewidth=2,
                    color="#F58518",
                )
                axes[1].set_title(f"按{period_column[1] if period_column else '期间'}趋势")
                axes[1].set_xlabel("期间")
                axes[1].set_ylabel(value_name)
                axes[1].tick_params(axis="x", rotation=35)
                axes[1].grid(True, alpha=0.25)
            else:
                sample = numeric_values[:20]
                axes[1].plot(range(1, len(sample) + 1), sample, marker="o", linewidth=2)
                axes[1].set_title(f"{value_name}前 {len(sample)} 条记录")
                axes[1].set_xlabel("记录序号")
                axes[1].set_ylabel(value_name)
                axes[1].grid(True, alpha=0.25)
            fig.tight_layout(rect=(0, 0, 1, 0.93))
            artifact_root = (
                Path(self._workspace_root or tempfile.gettempdir())
                / ".koto_artifacts"
                / "financial_charts"
            )
            artifact_root.mkdir(parents=True, exist_ok=True)
            chart_path = artifact_root / f"supplemental_chart_{uuid.uuid4().hex[:10]}.png"
            fig.savefig(chart_path, dpi=220, bbox_inches="tight")
            plt.close(fig)

            total_value = sum(numeric_values)
            top_label = top_categories[0][0] if top_categories else ""
            top_value = top_categories[0][1] if top_categories else None
            paragraphs: List[Dict[str, str]] = [
                {"text": f"{display_label}补充分析", "style": "Heading 2"},
                {
                    "text": f"数据来源：{display_label} / {sheet_name}；识别到 {len(records)} 条数据记录，核心数值列为“{value_name}”。"
                },
                {
                    "text": f"“{value_name}”合计约 {total_value:,.2f}；需要和主财务模型中的收入、销量或回款预测口径交叉核对。",
                    "style": "List Bullet",
                },
            ]
            if top_label and top_value is not None:
                paragraphs.append(
                    {
                        "text": f"最高贡献项为“{top_label}”，对应 {value_name} 约 {top_value:,.2f}。",
                        "style": "List Bullet",
                    }
                )
            if len(sorted_periods) >= 2:
                paragraphs.append(
                    {
                        "text": f"已按“{period_column[1] if period_column else '期间'}”生成趋势图，可用于核对销售节奏是否支撑财务预测。",
                        "style": "List Bullet",
                    }
                )
            else:
                paragraphs.append(
                    {
                        "text": "未识别到稳定期间列，图表使用记录序列展示，建议补充月份或日期字段后再做预测匹配。",
                        "style": "List Bullet",
                    }
                )
            return {
                "success": True,
                "path": str(chart_path),
                "sheet": sheet_name,
                "value_column": value_name,
                "category_column": category_column[1] if category_column else "",
                "period_column": period_column[1] if period_column else "",
                "record_count": len(records),
                "total_value": total_value,
                "top_categories": top_categories,
                "period_totals": sorted_periods,
                "paragraphs": paragraphs,
                "summary": f"已从“{sheet_name}”生成补充 Excel 图表：{chart_path.name}",
                "caption": f"数据来源：{display_label} / {sheet_name}；按 {value_name} 汇总展示分类贡献和期间趋势。",
            }
        except Exception as exc:
            return {"success": False, "error": str(exc)}

    def _generate_supplemental_xlsx_chart_pil(
        self,
        *,
        xlsx_path: str,
        sheet_name: str,
        value_name: str,
        category_name: str,
        period_name: str,
        record_count: int,
        numeric_values: List[float],
        top_categories: List[tuple[str, float]],
        sorted_periods: List[tuple[str, float]],
        display_name: str = "",
    ) -> Dict[str, Any]:
        display_label = str(display_name or self._display_path(xlsx_path)).strip()
        artifact_root = (
            Path(self._workspace_root or tempfile.gettempdir())
            / ".koto_artifacts"
            / "financial_charts"
        )
        chart_path = artifact_root / f"supplemental_chart_{uuid.uuid4().hex[:10]}.png"
        if len(sorted_periods) >= 2:
            trend_labels = [item[0] for item in sorted_periods]
            trend_values = [item[1] for item in sorted_periods]
        else:
            sample = numeric_values[:20]
            trend_labels = [str(index + 1) for index in range(len(sample))]
            trend_values = sample
        panels = [
            {
                "kind": "bar",
                "title": f"By {category_name or 'category'}",
                "bars": top_categories,
            },
            {
                "kind": "line",
                "title": f"{value_name} trend",
                "labels": trend_labels,
                "series": {value_name: trend_values},
            },
        ]
        self._save_pil_chart_image(
            chart_path,
            title=f"{display_label} Supplemental Analysis",
            panels=panels,
        )
        total_value = sum(numeric_values)
        top_label = top_categories[0][0] if top_categories else ""
        top_value = top_categories[0][1] if top_categories else None
        paragraphs: List[Dict[str, str]] = [
            {"text": f"{display_label}补充分析", "style": "Heading 2"},
            {
                "text": f"数据来源：{display_label} / {sheet_name}；识别到 {record_count} 条数据记录，核心数值列为“{value_name}”。"
            },
            {
                "text": f"“{value_name}”合计约 {total_value:,.2f}；需要和主财务模型中的收入、销量或回款预测口径交叉核对。",
                "style": "List Bullet",
            },
        ]
        if top_label and top_value is not None:
            paragraphs.append(
                {
                    "text": f"最高贡献项为“{top_label}”，对应 {value_name} 约 {top_value:,.2f}。",
                    "style": "List Bullet",
                }
            )
        if len(sorted_periods) >= 2:
            paragraphs.append(
                {
                    "text": f"已按“{period_name or '期间'}”生成趋势图，可用于核对销售节奏是否支撑财务预测。",
                    "style": "List Bullet",
                }
            )
        else:
            paragraphs.append(
                {
                    "text": "未识别到稳定期间列，图表使用记录序列展示，建议补充月份或日期字段后再做预测匹配。",
                    "style": "List Bullet",
                }
            )
        return {
            "success": True,
            "path": str(chart_path),
            "sheet": sheet_name,
            "value_column": value_name,
            "category_column": category_name,
            "period_column": period_name,
            "record_count": record_count,
            "total_value": total_value,
            "top_categories": top_categories,
            "period_totals": sorted_periods,
            "paragraphs": paragraphs,
            "summary": f"已使用 Pillow 兜底生成补充 Excel 图表：{chart_path.name}",
            "caption": f"数据来源：{display_label} / {sheet_name}；按 {value_name} 汇总展示分类贡献和期间趋势。",
        }

    def _financial_series_issues(
        self, series: Dict[str, List[Optional[float]]], years: List[str]
    ) -> List[str]:
        issues: List[str] = []
        for name, values in series.items():
            missing = [
                years[idx]
                for idx, value in enumerate(values)
                if value is None and idx < len(years)
            ]
            if missing:
                issues.append(
                    f"{name} 在 {', '.join(missing[:4])} 缺少有效数据，图表和结论需要回到底稿核对。"
                )
            previous: Optional[float] = None
            for idx, value in enumerate(values):
                if value is None:
                    continue
                if previous not in (None, 0):
                    growth = (value - previous) / abs(previous)
                    if growth > 1.0:
                        issues.append(
                            f"{name} 在 {years[idx] if idx < len(years) else '后续年份'} 同比增长超过 100%，假设偏激进，需要补充驱动解释。"
                        )
                    elif growth < -0.5:
                        issues.append(
                            f"{name} 在 {years[idx] if idx < len(years) else '后续年份'} 同比下滑超过 50%，需要确认是否为模型口径变化或录入问题。"
                        )
                previous = value
        deduped: List[str] = []
        seen: set[str] = set()
        for issue in issues:
            if issue in seen:
                continue
            seen.add(issue)
            deduped.append(issue)
        return deduped[:8]

    def _financial_series_movements(
        self, series: Dict[str, List[Optional[float]]], years: List[str]
    ) -> List[str]:
        movements: List[str] = []
        for name, values in series.items():
            valid = [
                (idx, value)
                for idx, value in enumerate(values)
                if value is not None and idx < len(years)
            ]
            if len(valid) < 2:
                continue
            first_idx, first_value = valid[0]
            last_idx, last_value = valid[-1]
            is_rate = bool(re.search(r"(?:率|增速|%|margin)", name, re.IGNORECASE))
            first_number = float(first_value)
            last_number = float(last_value)
            if is_rate:
                first_text = f"{first_number:.1%}"
                last_text = f"{last_number:.1%}"
                change_text = f"变化 {((last_number - first_number) * 100):.1f} 个百分点"
            elif first_number < 0 < last_number:
                first_text = f"{first_number:,.2f}"
                last_text = f"{last_number:,.2f}"
                change_text = "由负转正，累计增幅口径不适用"
            elif first_number == 0:
                first_text = f"{first_number:,.2f}"
                last_text = f"{last_number:,.2f}"
                change_text = "期初值为 0，无法计算累计增幅"
            else:
                first_text = f"{first_number:,.2f}"
                last_text = f"{last_number:,.2f}"
                growth = (last_number - first_number) / abs(first_number)
                change_text = f"累计变化 {growth:.1%}"
            movements.append(
                f"{name}：{years[first_idx]} 为 {first_text}，"
                f"{years[last_idx]} 为 {last_text}，{change_text}。"
            )
            if len(movements) >= 8:
                break
        return movements

    def _financial_report_model_synthesis(
        self,
        request: FileTaskRequest,
        audit_payload: Optional[Dict[str, Any]],
        inspect_payload: Optional[Dict[str, Any]],
        chart_result: Dict[str, Any],
    ) -> str:
        options = request.options if isinstance(request.options, dict) else {}
        if options.get("disable_financial_model_synthesis") is True:
            return ""
        facts = {
            "task": request.task,
            "workbook_summary": (
                (inspect_payload or {}).get("summary")
                if isinstance(inspect_payload, dict)
                else ""
            ),
            "audit_summary": (
                (audit_payload or {}).get("summary")
                if isinstance(audit_payload, dict)
                else ""
            ),
            "audit_findings": (
                (audit_payload or {}).get("findings", [])[:12]
                if isinstance(audit_payload, dict)
                else []
            ),
            "external_link_count": (
                (inspect_payload or {}).get("external_link_count")
                if isinstance(inspect_payload, dict)
                else None
            ),
            "total_formula_cells": (
                (inspect_payload or {}).get("total_formula_cells")
                if isinstance(inspect_payload, dict)
                else None
            ),
            "chart_sheet": chart_result.get("sheet"),
            "chart_years": chart_result.get("years"),
            "chart_series": chart_result.get("series"),
            "chart_series_groups": chart_result.get("series_groups"),
            "chart_issues": chart_result.get("issues"),
        }
        prompt = (
            "请基于以下 Excel 财务模型审计事实，输出可直接写入 Word 的中文问题分析。"
            "要求：不要编造未给出的事实；优先指出模型可靠性、假设激进性、公式/外链/口径风险；"
            "用 4-8 条短要点，每条一句话。\n\n"
            f"{json.dumps(facts, ensure_ascii=False, default=str)[:6000]}"
        )
        try:
            response = self._call_model(
                request=request,
                messages=[{"role": "user", "content": prompt}],
                system="你是严谨的财务模型审阅助手。只输出分析要点，不调用工具，不声称已修改文件。",
                tools=[],
            )
            content, _ = self._normalize_model_response(response, [])
        except Exception as exc:
            logger.info("[FileTaskRuntime] financial model synthesis skipped: %s", exc)
            return ""
        return _preview(content, 1800)

    def _merge_financial_model_synthesis(
        self, paragraphs: List[Dict[str, str]], model_synthesis: str
    ) -> List[Dict[str, str]]:
        clean_lines: List[str] = []
        for raw_line in str(model_synthesis or "").splitlines():
            line = re.sub(r"^\s*(?:[-*•]|\d+[.)、])\s*", "", raw_line).strip()
            line = re.sub(r"^#+\s*", "", line).strip()
            if not line or line in clean_lines:
                continue
            clean_lines.append(line)
            if len(clean_lines) >= 8:
                break
        if not clean_lines:
            return paragraphs
        insert_at = len(paragraphs)
        for idx, item in enumerate(paragraphs):
            if str(item.get("text") or "").strip() == "图表说明":
                insert_at = idx
                break
        supplement: List[Dict[str, str]] = [{"text": "AI 综合分析", "style": "Heading 2"}]
        supplement.extend(
            {"text": line, "style": "List Bullet"} for line in clean_lines
        )
        return paragraphs[:insert_at] + supplement + paragraphs[insert_at:]

    def _financial_latest_ratio(
        self, numerator: List[Optional[float]], denominator: List[Optional[float]]
    ) -> Optional[float]:
        for num, den in zip(reversed(numerator), reversed(denominator)):
            if num is None or den in (None, 0):
                continue
            return float(num) / abs(float(den))
        return None

    def _financial_format_metric(
        self, value: Optional[float], *, percent: bool = False
    ) -> str:
        if value is None:
            return "缺失"
        if percent:
            return f"{float(value):.1%}"
        return f"{float(value):,.2f}"

    def _financial_growth_between(
        self, values: List[Optional[float]]
    ) -> Optional[float]:
        valid = [float(value) for value in values if value is not None]
        if len(valid) < 2 or valid[0] == 0:
            return None
        return (valid[-1] - valid[0]) / abs(valid[0])

    def _financial_report_executive_summary(
        self, chart_result: Dict[str, Any]
    ) -> List[str]:
        groups = (
            chart_result.get("series_groups")
            if isinstance(chart_result.get("series_groups"), dict)
            else {}
        )
        money = groups.get("money") if isinstance(groups.get("money"), dict) else {}
        rates = groups.get("rates") if isinstance(groups.get("rates"), dict) else {}
        expenses = (
            groups.get("expenses") if isinstance(groups.get("expenses"), dict) else {}
        )
        volume = groups.get("volume") if isinstance(groups.get("volume"), dict) else {}
        lines: List[str] = []
        revenue = money.get("收入合计") or []
        net_profit = money.get("净利润") or []
        gross_profit = money.get("毛利合计") or []
        if revenue:
            lines.append(
                f"收入预测期累计变化为 {self._financial_format_metric(self._financial_growth_between(revenue), percent=True)}，属于本模型最核心的增长假设。"
            )
        if net_profit and revenue:
            lines.append(
                f"末期净利率约 {self._financial_format_metric(self._financial_latest_ratio(net_profit, revenue), percent=True)}，需与费用率、毛利率假设联动核对。"
            )
        if gross_profit and revenue:
            lines.append(
                f"末期毛利率约 {self._financial_format_metric(self._financial_latest_ratio(gross_profit, revenue), percent=True)}，需确认硬件、配件和互联网业务口径是否一致。"
            )
        if volume.get("总销量"):
            lines.append(
                f"销量预测期累计变化为 {self._financial_format_metric(self._financial_growth_between(volume.get('总销量') or []), percent=True)}，需要拆解到产品线、渠道和产能约束。"
            )
        if expenses.get("费用合计") and revenue:
            lines.append(
                f"末期费用率约 {self._financial_format_metric(self._financial_latest_ratio(expenses.get('费用合计') or [], revenue), percent=True)}，需要和市场投放、研发团队扩张节奏匹配。"
            )
        return lines[:6] or ["已识别关键财务预测指标，但仍需补充底层假设说明后才能形成投资判断。"]

    def _financial_assumption_risks(self, chart_result: Dict[str, Any]) -> List[str]:
        groups = (
            chart_result.get("series_groups")
            if isinstance(chart_result.get("series_groups"), dict)
            else {}
        )
        money = groups.get("money") if isinstance(groups.get("money"), dict) else {}
        rates = groups.get("rates") if isinstance(groups.get("rates"), dict) else {}
        volume = groups.get("volume") if isinstance(groups.get("volume"), dict) else {}
        risks: List[str] = []
        revenue_growth = self._financial_growth_between(money.get("收入合计") or [])
        volume_growth = self._financial_growth_between(volume.get("总销量") or [])
        net_margin = rates.get("净利率") or []
        gross_margin = rates.get("综合毛利率") or []
        if revenue_growth is not None and revenue_growth > 3:
            risks.append("收入预测期累计增长超过 300%，需要把增长拆到销量、ASP、产品结构和区域扩张，不能只停留在结果行。")
        if volume_growth is not None and volume_growth > 3:
            risks.append("销量预测期放量幅度很大，需要补充产能、渠道、价格带和竞品压力的约束条件。")
        if net_margin and any(
            value is not None and value > 0.1 for value in net_margin
        ):
            risks.append("净利率在预测期进入较高区间，需要核查销售费用率和研发费用率是否被过早摊薄。")
        if (
            gross_margin
            and len(
                {round(float(value), 4) for value in gross_margin if value is not None}
            )
            <= 2
        ):
            risks.append("毛利率曲线变化较少，可能存在硬编码或未充分反映产品结构变化。")
        risks.extend(
            str(item)
            for item in chart_result.get("issues") or []
            if str(item or "").strip()
        )
        deduped: List[str] = []
        for item in risks:
            if item not in deduped:
                deduped.append(item)
        return deduped[:8]

    def _financial_followup_questions(self, chart_result: Dict[str, Any]) -> List[str]:
        return [
            "收入增长的核心驱动是销量、ASP、产品结构还是海外市场扩张？每一项分别贡献多少？",
            "XR、AI、AR 各产品线的销量假设对应哪些渠道、价格带和竞品对标？",
            "毛利率改善来自规模效应、供应链降本、产品组合变化，还是互联网业务占比提升？",
            "销售费用率是否充分反映新品上市、达人投放、海外渠道建设和退换货成本？",
            "研发费用率下降是否和团队招聘、芯片/光学/算法投入计划一致？",
            "外部链接对应哪些底稿？如果缺失，哪些关键输出无法复算？",
        ]

    def _financial_report_problem_paragraphs(
        self,
        audit_payload: Optional[Dict[str, Any]],
        inspect_payload: Optional[Dict[str, Any]],
        chart_result: Dict[str, Any],
    ) -> List[Dict[str, str]]:
        paragraphs: List[Dict[str, str]] = [
            {"text": "财务模型分析图表与问题", "style": "Heading 1"},
            {"text": str(chart_result.get("caption") or "已基于附件 Excel 财务模型生成图表和问题清单。")},
            {"text": "核心结论", "style": "Heading 2"},
        ]
        paragraphs.extend(
            {"text": item, "style": "List Bullet"}
            for item in self._financial_report_executive_summary(chart_result)
        )
        paragraphs.extend(
            [
                {"text": "数据口径", "style": "Heading 2"},
                {
                    "text": f"图表取数工作表：{chart_result.get('sheet') or '未识别'}；年份列：{', '.join(chart_result.get('years') or []) or '未识别'}。"
                },
            ]
        )
        series = (
            chart_result.get("series")
            if isinstance(chart_result.get("series"), dict)
            else {}
        )
        movements = self._financial_series_movements(
            series, chart_result.get("years") or []
        )
        if movements:
            paragraphs.append({"text": "关键指标变化", "style": "Heading 2"})
            paragraphs.extend(
                {"text": item, "style": "List Bullet"} for item in movements
            )
        assumption_risks = self._financial_assumption_risks(chart_result)
        if assumption_risks:
            paragraphs.append({"text": "经营假设风险", "style": "Heading 2"})
            paragraphs.extend(
                {"text": item, "style": "List Bullet"} for item in assumption_risks
            )
        paragraphs.extend(
            [
                {"text": "模型质量问题", "style": "Heading 2"},
            ]
        )
        issues: List[str] = []
        audit = audit_payload if isinstance(audit_payload, dict) else {}
        inspect_data = inspect_payload if isinstance(inspect_payload, dict) else {}
        for finding in audit.get("findings") or []:
            if isinstance(finding, dict):
                severity = str(finding.get("severity") or "").strip()
                message = str(finding.get("message") or "").strip()
                location = str(
                    finding.get("location") or finding.get("sheet") or ""
                ).strip()
                if message:
                    suffix = f"（位置：{location}）" if location else ""
                    issues.append(f"[{severity or 'info'}] {message}{suffix}")
        external_count = inspect_data.get("external_link_count")
        if external_count:
            issues.append(f"工作簿检测到 {external_count} 个外部链接，模型复算依赖外部文件，需补齐底稿或解除外链。")
        formula_count = inspect_data.get("total_formula_cells")
        if formula_count:
            issues.append(f"工作簿共检测到 {formula_count} 个公式单元格，建议重点核查关键输出行的公式连续性。")
        issues.extend(
            str(item)
            for item in chart_result.get("issues") or []
            if str(item or "").strip()
        )
        if not chart_result.get("success"):
            issues.append(f"图表生成未完全成功：{chart_result.get('error') or '缺少可作图数据'}。")
        if not issues:
            issues.append("未发现明显结构性红旗；仍建议核对关键假设、外部链接和历史口径。")

        seen: set[str] = set()
        for issue in issues:
            text = str(issue or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            paragraphs.append({"text": text, "style": "List Bullet"})
            if len(paragraphs) >= 34:
                break
        paragraphs.append({"text": "建议追问", "style": "Heading 2"})
        paragraphs.extend(
            {"text": item, "style": "List Bullet"}
            for item in self._financial_followup_questions(chart_result)
        )
        paragraphs.append({"text": "图表说明", "style": "Heading 2"})
        paragraphs.append({"text": chart_result.get("summary") or "图表已按模型中的关键指标生成。"})
        return paragraphs


    def _emit_chart_pipeline_events(
        self,
        ledger: FileTaskLedger,
        *,
        step_id: str,
        code: str,
        result: Dict[str, Any],
        default_message: str,
        source_path: str = "",
    ) -> Iterable[FileTaskEvent]:
        code_payload: Dict[str, Any] = {"code": code}
        if source_path:
            code_payload["source_path"] = source_path
        yield ledger.event("code.started", code_payload, step_id=step_id)
        yield ledger.event(
            "code.output",
            {
                "text": result.get("summary")
                or result.get("error")
                or default_message,
                "stream": "stdout" if result.get("success") else "stderr",
            },
            step_id=step_id,
        )
        yield ledger.event(
            "code.finished",
            {"success": bool(result.get("success"))},
            step_id=step_id,
        )
        yield ledger.event(
            "tool.finished",
            {
                "tool_name": "run_python_code",
                "success": bool(result.get("success")),
                "result_preview": result.get("summary")
                or result.get("error")
                or default_message,
            },
            step_id=step_id,
        )
