# -*- coding: utf-8 -*-
"""Deterministic recovery for the financial XLSX -> charted DOCX recipe.

This is not a second routing path.  The whitebox mainline invokes it only when
the model finishes without producing a file change.  Every mutation still goes
through the canonical allowlisted tools and emits the normal tool/file events.
"""

from __future__ import annotations

from dataclasses import dataclass, field
import json
import logging
import math
import os
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

from app.core.agent.file_task_contract import (
    FileTaskEvent,
    FileTaskToolStreamResult,
)
from app.core.agent.file_task_artifact_transaction import (
    cleanup_run_owned_paths,
    commit_staged_artifact,
    committed_file_changes,
    run_scoped_staging_path,
)
from app.core.agent.file_task_runtime_utils import _is_error_result
from app.core.agent.file_task_tool_catalog import tool_result_preview


logger = logging.getLogger(__name__)
_IMAGE_ARTIFACT_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "bmp", "gif"}


@dataclass
class FinancialReportRecoveryResult:
    attempted: bool = False
    completed: bool = False
    summary: str = ""
    file_changes: List[Dict[str, Any]] = field(default_factory=list)
    artifacts: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class _ToolOutcome:
    success: bool
    payload: Dict[str, Any]
    result: Any
    changes: List[Dict[str, Any]]
    artifacts: List[Dict[str, Any]]


def _json_mapping(value: Any) -> Dict[str, Any]:
    if isinstance(value, dict):
        return dict(value)
    try:
        parsed = json.loads(str(value or ""))
    except Exception:
        return {}
    return dict(parsed) if isinstance(parsed, dict) else {}


def _run_tool(
    runtime: Any,
    ledger: Any,
    executor: Any,
    *,
    step_id: str,
    tool_name: str,
    tool_args: Dict[str, Any],
    event_metadata: Optional[Dict[str, Any]] = None,
    emit_file_changes: bool = True,
) -> Iterable[FileTaskEvent | _ToolOutcome]:
    metadata = dict(event_metadata or {})
    yield ledger.event(
        "tool.started",
        {"tool_name": tool_name, "tool_args": dict(tool_args), **metadata},
        step_id=step_id,
    )
    try:
        result = executor(tool_name, dict(tool_args))
        if isinstance(result, FileTaskToolStreamResult):
            result = yield from runtime._consume_streaming_tool_result(
                ledger,
                step_id=step_id,
                stream_result=result,
            )
        success = not _is_error_result(result)
    except Exception as exc:
        result = {"error": str(exc)}
        success = False

    yield ledger.event(
        "tool.finished",
        {
            "tool_name": tool_name,
            "success": success,
            "result_preview": tool_result_preview(tool_name, result, 1200),
            **metadata,
        },
        step_id=step_id,
    )
    changes = (
        runtime._extract_file_changes(tool_name, tool_args, result) if success else []
    )
    if emit_file_changes:
        for change in changes:
            yield ledger.event("file.changed", change, step_id=step_id)
    artifacts = runtime._tool_artifacts(tool_name, result) if success else []
    return _ToolOutcome(
        success=success,
        payload=_json_mapping(result),
        result=result,
        changes=changes,
        artifacts=artifacts,
    )


def _expected_financial_chart_paths(source_path: str) -> List[Path]:
    source = Path(source_path)
    stem = (
        re.sub(r"[^A-Za-z0-9_-]+", "_", source.stem).strip("_")
        or "financial_model"
    )
    return [
        source.parent / f"{stem}_financial_pnl_trend.png",
        source.parent / f"{stem}_product_sales_revenue_structure.png",
    ]


def _file_type(file_info: Any) -> str:
    raw = str(getattr(file_info, "type", "") or "").strip().lower().lstrip(".")
    if raw:
        return raw
    return Path(str(getattr(file_info, "path", "") or "")).suffix.lower().lstrip(".")


def _source_workbook(runtime: Any, files: List[Any]) -> str:
    candidates = [
        str(getattr(item, "path", "") or "").strip()
        for item in files
        if _file_type(item) in {"xlsx", "xlsm", "xls"}
        and not bool(getattr(item, "target", False))
    ]
    for candidate in candidates:
        resolved = runtime._resolve_task_file_path(candidate)
        if resolved:
            return str(resolved)
    return ""


def _as_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace(",", "")
    if not text or text.startswith("="):
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _extract_series(source_path: str) -> Tuple[str, List[str], Dict[str, List[Optional[float]]]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return "", [], {}

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    try:
        sheet = next(
            (
                item
                for item in workbook.worksheets
                if re.search(r"(?:p\s*&\s*l|利润|损益|income)", item.title, re.I)
            ),
            workbook.worksheets[0] if workbook.worksheets else None,
        )
        if sheet is None:
            return "", [], {}
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        header_index = -1
        year_columns: List[Tuple[int, str]] = []
        for index, row in enumerate(rows[:30]):
            matches = []
            for column, value in enumerate(row):
                text = str(value or "").strip()
                if re.fullmatch(r"20\d{2}[AE]?", text, re.I):
                    matches.append((column, text))
            if len(matches) >= 2:
                header_index = index
                year_columns = matches
                break
        if header_index < 0:
            return sheet.title, [], {}

        series: Dict[str, List[Optional[float]]] = {}
        for row in rows[header_index + 1 :]:
            label = next(
                (
                    str(value).strip()
                    for value in row[: year_columns[0][0] + 1]
                    if isinstance(value, str) and str(value).strip()
                ),
                "",
            )
            if not label:
                continue
            values = [
                _as_number(row[column] if column < len(row) else None)
                for column, _ in year_columns
            ]
            if any(value is not None for value in values):
                series.setdefault(label, values)
        return sheet.title, [label for _, label in year_columns], series
    finally:
        workbook.close()


def _pick_series(
    series: Dict[str, List[Optional[float]]], *patterns: str
) -> Tuple[str, List[Optional[float]]]:
    for label, values in series.items():
        lowered = label.lower()
        if any(pattern.lower() in lowered for pattern in patterns):
            return label, values
    return "", []


def _first_last(values: List[Optional[float]]) -> Tuple[Optional[float], Optional[float]]:
    present = [value for value in values if value is not None]
    return (present[0], present[-1]) if present else (None, None)


def _number(value: Optional[float]) -> str:
    return "—" if value is None else f"{value:,.1f}"


def _percent(value: Optional[float]) -> str:
    if value is None:
        return "—"
    ratio = value if abs(value) <= 2 else value / 100.0
    return f"{ratio * 100:.1f}%"


def _build_report_paragraphs(
    source_path: str,
    audit: Dict[str, Any],
) -> Tuple[List[Dict[str, str]], str]:
    sheet_name, periods, series = _extract_series(source_path)
    revenue_label, revenue = _pick_series(series, "收入合计", "营业收入", "revenue")
    gross_label, gross_profit = _pick_series(series, "毛利合计", "毛利", "gross profit")
    gross_margin_label, gross_margin = _pick_series(series, "综合毛利率", "毛利率", "gross margin")
    net_label, net_profit = _pick_series(series, "净利润", "net profit")
    net_margin_label, net_margin = _pick_series(series, "净利率", "net margin")
    tax_label, tax = _pick_series(series, "所得税费用", "income tax")

    revenue_first, revenue_last = _first_last(revenue)
    gross_first, gross_last = _first_last(gross_profit)
    net_first, net_last = _first_last(net_profit)
    period_count = max(1, len([item for item in revenue if item is not None]) - 1)
    revenue_cagr = None
    if revenue_first and revenue_last is not None and revenue_first > 0 and revenue_last >= 0:
        revenue_cagr = (revenue_last / revenue_first) ** (1 / period_count) - 1

    findings = [
        str(item.get("message") or "").strip()
        for item in audit.get("findings") or []
        if isinstance(item, dict) and str(item.get("message") or "").strip()
    ]
    if revenue_cagr is not None and revenue_cagr > 0.5:
        findings.append(
            f"收入预测复合增速约为 {_percent(revenue_cagr)}，增长假设显著偏激进，需补充销量、单价和渠道渗透率的敏感性分析。"
        )
    net_values = [value for value in net_profit if value is not None]
    if net_values and min(net_values) < 0 < max(net_values):
        findings.append("预测期内净利润由亏损转为盈利，盈利拐点高度依赖收入增长和费用率下降，建议建立下行情景。")
    if tax and any(value is None for value in tax):
        findings.append("所得税费用预测存在年度空值，税率、亏损弥补和递延所得税逻辑需要逐年闭环。")
    if tax and any((value or 0) < 0 for value in tax if value is not None):
        findings.append("所得税费用出现负值，应明确其属于税收返还、亏损抵扣还是符号错误，并与净利润公式复核。")
    if gross_margin:
        gm_first, gm_last = _first_last(gross_margin)
        if gm_first is not None and gm_last is not None and gm_last - gm_first > 0.08:
            findings.append(
                f"综合毛利率由 {_percent(gm_first)} 上升至 {_percent(gm_last)}，改善幅度较大，需要用产品结构、采购降本和规模效应逐项支撑。"
            )
    unique_findings: List[str] = []
    for item in findings:
        if item and item not in unique_findings:
            unique_findings.append(item)
    if not unique_findings:
        unique_findings.append("未发现可直接判定的结构性错误，但预测模型仍应补充关键假设、敏感性分析和三表勾稽检查。")

    period_text = "、".join(periods) if periods else "可用预测年度"
    summary = (
        f"模型覆盖 {period_text}；{revenue_label or '收入'}由 {_number(revenue_first)} 增至 {_number(revenue_last)}，"
        f"{net_label or '净利润'}由 {_number(net_first)} 变为 {_number(net_last)}。"
    )
    paragraphs: List[Dict[str, str]] = [
        {"text": "财务预测分析报告", "style": "Title"},
        {
            "text": f"数据来源：{Path(source_path).name}；核心分析工作表：{sheet_name or '未识别'}。本报告由 Koto 原生恢复流程生成，所有结论均基于工作簿可读取数据。",
            "style": "Normal",
        },
        {"text": "一、核心结论", "style": "Heading 1"},
        {"text": summary, "style": "Normal"},
        {"text": "二、关键财务趋势", "style": "Heading 1"},
        {
            "text": f"收入趋势：{revenue_label or '收入指标'}从 {_number(revenue_first)} 变为 {_number(revenue_last)}"
            + (f"，预测期复合增速约 {_percent(revenue_cagr)}。" if revenue_cagr is not None else "。"),
            "style": "Normal",
        },
        {
            "text": f"毛利趋势：{gross_label or '毛利指标'}从 {_number(gross_first)} 变为 {_number(gross_last)}。",
            "style": "Normal",
        },
        {
            "text": f"盈利趋势：{net_label or '净利润指标'}从 {_number(net_first)} 变为 {_number(net_last)}；"
            f"{net_margin_label or '净利率'}最新值为 {_percent(_first_last(net_margin)[1])}。",
            "style": "Normal",
        },
        {"text": "三、发现的问题与风险", "style": "Heading 1"},
    ]
    for index, finding in enumerate(unique_findings[:12], start=1):
        paragraphs.append(
            {"text": f"问题 {index}：{finding}", "style": "Heading 2"}
        )
    paragraphs.extend(
        [
            {"text": "四、建议的补充验证", "style": "Heading 1"},
            {
                "text": "建议至少建立基准、下行和压力三种情景，并分别校验收入、毛利率、费用率、现金消耗和融资需求。",
                "style": "Normal",
            },
            {
                "text": "建议补齐资产负债表与现金流量表的逐年勾稽，并消除外部链接或附带完整底稿后再用于正式决策。",
                "style": "Normal",
            },
        ]
    )
    return paragraphs, sheet_name


def recover_financial_report(
    runtime: Any,
    ledger: Any,
    request: Any,
    executor: Any,
    files: List[Any],
    recipe_skeleton: Dict[str, Any],
    *,
    step_id: str,
) -> Iterable[FileTaskEvent | FinancialReportRecoveryResult]:
    recipe_id = str((recipe_skeleton or {}).get("recipe_id") or "").strip()
    if recipe_id != "financial_xlsx_docx_report":
        return FinancialReportRecoveryResult()

    source_path = _source_workbook(runtime, files)
    target_path = str(getattr(request, "target_path", "") or "").strip()
    if not source_path or not target_path:
        return FinancialReportRecoveryResult(
            attempted=True,
            summary="财务报告原生恢复无法启动：源工作簿或目标 DOCX 路径不明确。",
        )

    staging_path = run_scoped_staging_path(request, target_path)
    expected_chart_paths = _expected_financial_chart_paths(source_path)
    preexisting_chart_paths = {
        path for path in expected_chart_paths if path.is_file()
    }
    cleanup_run_owned_paths(staging_path)

    yield ledger.event(
        "recovery.started",
        {
            "kind": "financial_report",
            "summary": "模型未完成写入，正在使用 Koto 原生财务报告流程恢复任务。",
        },
        step_id=step_id,
    )
    all_changes: List[Dict[str, Any]] = []
    all_artifacts: List[Dict[str, Any]] = []

    audit_outcome = yield from _run_tool(
        runtime,
        ledger,
        executor,
        step_id=step_id,
        tool_name="audit_financial_workbook",
        tool_args={"path": source_path, "max_findings": 16},
        event_metadata={"native_financial_recovery": True},
        emit_file_changes=False,
    )
    if not audit_outcome.success:
        return FinancialReportRecoveryResult(
            attempted=True,
            summary="财务报告原生恢复失败：无法审计源工作簿。",
        )

    paragraphs, sheet_name = _build_report_paragraphs(
        source_path, audit_outcome.payload
    )
    chart_args = runtime._financial_chart_recovery_tool_args(
        request,
        files,
        {"code": "import matplotlib.pyplot as plt\nplt.savefig('financial_chart.png')"},
        [],
    )
    if chart_args:
        chart_outcome = yield from _run_tool(
            runtime,
            ledger,
            executor,
            step_id=step_id,
            tool_name="run_python_code",
            tool_args=chart_args,
            event_metadata={"native_financial_recovery": True},
            emit_file_changes=False,
        )
    else:
        chart_outcome = _ToolOutcome(False, {}, {}, [], [])
    all_changes.extend(chart_outcome.changes)
    all_artifacts.extend(chart_outcome.artifacts)

    write_outcome = yield from _run_tool(
        runtime,
        ledger,
        executor,
        step_id=step_id,
        tool_name="write_docx_content",
        tool_args={
            "path": str(staging_path),
            "paragraphs": json.dumps(paragraphs, ensure_ascii=False),
        },
        event_metadata={"native_financial_recovery": True},
        emit_file_changes=False,
    )
    all_changes.extend(write_outcome.changes)

    table_args: Dict[str, Any] = {
        "source_path": source_path,
        "target_path": str(staging_path),
        "table_title": "关键财务预测数据",
        "max_rows": 18,
        "financial_compact": True,
    }
    if sheet_name:
        table_args["sheet_name"] = sheet_name
    table_outcome = yield from _run_tool(
        runtime,
        ledger,
        executor,
        step_id=step_id,
        tool_name="insert_excel_as_docx_table",
        tool_args=table_args,
        event_metadata={"native_financial_recovery": True},
        emit_file_changes=False,
    )
    all_changes.extend(table_outcome.changes)

    image_paths: List[str] = []
    for artifact in chart_outcome.artifacts:
        if not isinstance(artifact, dict):
            continue
        path = str(artifact.get("path") or "").strip()
        if path and Path(path).suffix.lower() in {".png", ".jpg", ".jpeg"}:
            image_paths.append(path)
    if len(image_paths) < 2:
        image_paths.extend(
            str(path) for path in expected_chart_paths if path.exists()
        )
    image_paths = list(dict.fromkeys(image_paths))[:2]

    image_successes = 0
    for index, image_path in enumerate(image_paths, start=1):
        image_outcome = yield from _run_tool(
            runtime,
            ledger,
            executor,
            step_id=step_id,
            tool_name="insert_image_into_docx",
            tool_args={
                "path": str(staging_path),
                "image_path": image_path,
                "title": "财务趋势图" if index == 1 else "产品与销售结构图",
                "caption": f"图 {index}：基于源工作簿数据生成",
                "width_inches": 6.4,
            },
            event_metadata={"native_financial_recovery": True},
            emit_file_changes=False,
        )
        all_changes.extend(image_outcome.changes)
        if image_outcome.success:
            image_successes += 1

    tools_completed = bool(
        chart_outcome.success
        and write_outcome.success
        and table_outcome.success
        and image_successes >= 2
    )
    completed = bool(
        tools_completed and commit_staged_artifact(staging_path, target_path)
    )

    if completed:
        all_changes = committed_file_changes(
            all_changes,
            staging_path=staging_path,
            target_path=target_path,
        )
        for change in all_changes:
            yield ledger.event("file.changed", change, step_id=step_id)
    else:
        cleanup_run_owned_paths(
            staging_path,
            expected_chart_paths,
            preexisting_paths=preexisting_chart_paths,
        )
        all_changes = []
        all_artifacts = []

    summary = (
        "模型执行未完成，但 Koto 原生恢复流程已生成财务分析 DOCX、关键数据表和两张真实图表。"
        if completed
        else "财务报告原生恢复只完成了部分步骤，未通过完整产物核验。"
    )
    yield ledger.event(
        "recovery.finished",
        {
            "kind": "financial_report",
            "success": completed,
            "summary": summary,
            "file_change_count": len(all_changes),
            "image_count": image_successes,
        },
        step_id=step_id,
    )
    return FinancialReportRecoveryResult(
        attempted=True,
        completed=completed,
        summary=summary,
        file_changes=all_changes,
        artifacts=all_artifacts,
    )


def pending_generated_docx_images(
    runtime: Any,
    request: Any,
    files: List[Any],
    generated_artifacts: List[Dict[str, Any]],
    file_changes: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Return generated chart images that are not yet embedded in the DOCX."""
    if not generated_artifacts or not runtime._is_docx_chart_write_request(
        request, files
    ):
        return []

    inserted_keys: set[str] = set()
    inserted_names: set[str] = set()
    for change in file_changes or []:
        if not isinstance(change, dict):
            continue
        if str(change.get("operation") or "").strip() != "insert_image_into_docx":
            continue
        image_path = str(change.get("image_path") or "").strip()
        image_name = str(change.get("image_name") or "").strip()
        if image_path:
            inserted_keys.add(runtime._generated_image_artifact_key(image_path))
            inserted_names.add(Path(image_path.replace("\\", "/")).name.lower())
        if image_name:
            inserted_names.add(image_name.lower())

    pending: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for artifact in generated_artifacts or []:
        if not isinstance(artifact, dict):
            continue
        kind = str(artifact.get("kind") or "").strip().lower()
        path = str(artifact.get("path") or "").strip()
        name = runtime._generated_image_artifact_name(artifact)
        ext = Path(name or path).suffix.lstrip(".").lower()
        if kind != "image" or ext not in _IMAGE_ARTIFACT_EXTENSIONS:
            continue
        resolved_path = runtime._resolve_task_file_path(path) if path else ""
        if path and not (resolved_path or os.path.exists(path)):
            continue
        artifact = dict(artifact)
        if resolved_path:
            artifact["path"] = resolved_path
        key = runtime._generated_image_artifact_key(resolved_path or path or name)
        name_key = name.lower()
        if key in inserted_keys or name_key in inserted_names:
            continue
        if key in seen or name_key in seen:
            continue
        seen.add(key or name_key)
        pending.append(artifact)
    return pending


def financial_chart_recovery_tool_args(
    runtime: Any,
    request: Any,
    files: List[Any],
    original_tool_args: Dict[str, Any],
    artifacts: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Build verified financial charts when a model chart script produced none."""
    if artifacts or not runtime._looks_like_financial_xlsx_docx_chart_report_task(
        request, files
    ):
        return {}
    original_code = str((original_tool_args or {}).get("code") or "")
    if not re.search(
        r"(?:savefig|matplotlib|plt\.|chart|图表|绘图)",
        original_code,
        re.IGNORECASE,
    ):
        return {}
    source = runtime._single_source_path_for_types(request, files, {"xlsx", "xls"})
    source = runtime._resolve_task_file_path(source)
    if not source:
        return {}
    source_path = Path(source)
    output_dir = source_path.parent
    stem = re.sub(r"[^A-Za-z0-9_-]+", "_", source_path.stem).strip("_") or "financial_model"
    pnl_chart = output_dir / f"{stem}_financial_pnl_trend.png"
    structure_chart = output_dir / f"{stem}_product_sales_revenue_structure.png"
    return {
        "timeout": 120,
        "code": financial_chart_recovery_code(
            source_path, pnl_chart, structure_chart
        ),
    }


def financial_chart_recovery_code(
    source_path: Path, pnl_chart: Path, structure_chart: Path
) -> str:
    """Return self-contained, data-driven chart code for the recovery path."""
    return f'''from pathlib import Path
from openpyxl import load_workbook
import math
import re
import matplotlib.pyplot as plt

source_path = Path({json.dumps(str(source_path), ensure_ascii=False)})
pnl_chart = Path({json.dumps(str(pnl_chart), ensure_ascii=False)})
structure_chart = Path({json.dumps(str(structure_chart), ensure_ascii=False)})
pnl_chart.parent.mkdir(parents=True, exist_ok=True)

def as_number(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = str(value).strip().replace(',', '')
    if not text or text.startswith('='):
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None

workbook = load_workbook(source_path, data_only=True, read_only=True)
rows = []
for worksheet in workbook.worksheets:
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        label = next((str(item).strip() for item in values if isinstance(item, str) and item.strip()), '')
        numbers = [as_number(item) for item in values]
        numeric = [item for item in numbers if item is not None]
        if label and len(numeric) >= 2:
            rows.append((label, numeric))

if len(rows) < 2:
    raise RuntimeError('Unable to extract at least two numeric financial series from the workbook')

def pick_row(patterns, used):
    for index, (label, values) in enumerate(rows):
        lowered = label.lower()
        if index not in used and any(pattern in lowered for pattern in patterns):
            used.add(index)
            return label, values
    for index, item in enumerate(rows):
        if index not in used:
            used.add(index)
            return item
    return rows[0]

used = set()
revenue = pick_row(['revenue', '收入', 'sales'], used)
gross_profit = pick_row(['gross', '毛利'], used)
expenses = pick_row(['expense', '费用', '成本'], used)
net_profit = pick_row(['net', '净利', '利润'], used)
series = [revenue, gross_profit, expenses, net_profit]
period_count = min(len(values) for _, values in series)
periods = [f'P{{index + 1}}' for index in range(period_count)]

plt.style.use('seaborn-v0_8-whitegrid')
fig, axis = plt.subplots(figsize=(10, 5.6), dpi=160)
for label, values in series:
    axis.plot(periods, values[:period_count], marker='o', linewidth=2, label=label[:30])
axis.set_title('Financial P&L Trend')
axis.set_ylabel('Workbook value')
axis.legend(loc='best', fontsize=8)
fig.tight_layout()
fig.savefig(pnl_chart, bbox_inches='tight')
plt.close(fig)

product_rows = []
for index, item in enumerate(rows):
    if index in used:
        continue
    label, values = item
    lowered = label.lower()
    if any(token in lowered for token in ['product', '产品', '销量', 'volume', '销售', '收入']):
        product_rows.append(item)
    if len(product_rows) >= 6:
        break
if len(product_rows) < 2:
    product_rows = [item for index, item in enumerate(rows) if index not in used][:6]
if len(product_rows) < 2:
    product_rows = rows[:6]

fig, axis = plt.subplots(figsize=(10, 5.6), dpi=160)
labels = [label[:24] for label, _ in product_rows]
last_values = [values[-1] for _, values in product_rows]
bars = axis.bar(range(len(labels)), last_values, color='#4C78A8')
axis.set_xticks(range(len(labels)), labels, rotation=25, ha='right')
axis.set_title('Product / Sales Structure (Latest Available Period)')
axis.set_ylabel('Workbook value')
for bar, value in zip(bars, last_values):
    axis.annotate(f'{{value:,.0f}}', (bar.get_x() + bar.get_width() / 2, value),
                  xytext=(0, 3), textcoords='offset points', ha='center', va='bottom', fontsize=7)
fig.tight_layout()
fig.savefig(structure_chart, bbox_inches='tight')
plt.close(fig)

print('KOTO_CREATED:' + str(pnl_chart.resolve()))
print('KOTO_CREATED:' + str(structure_chart.resolve()))
'''


def insert_pending_generated_docx_images_native(
    runtime: Any,
    ledger: Any,
    request: Any,
    executor: Any,
    files: List[Any],
    pending_images: List[Dict[str, Any]],
    step_id: str,
) -> Iterable[Any]:
    """Deterministically insert already-created chart images into the target DOCX."""
    target = runtime._single_target_path_for_types(request, files, {"docx", "doc"})
    if not target:
        return []

    changes: List[Dict[str, Any]] = []
    for artifact in pending_images:
        image_path = str(artifact.get("path") or "").strip()
        if not image_path:
            continue
        image_name = runtime._generated_image_artifact_name(artifact)
        tool_args = {
            "path": target,
            "image_path": image_path,
            "title": "财务分析图表",
            "caption": image_name,
        }
        yield ledger.event(
            "tool.started",
            {
                "tool_name": "insert_image_into_docx",
                "tool_args": tool_args,
                "native_chart_completion": True,
            },
            step_id=step_id,
        )
        try:
            result = executor("insert_image_into_docx", tool_args)
            success = not _is_error_result(result)
        except Exception as exc:
            result = f"Error: {exc}"
            success = False
            logger.warning(
                "native chart insertion failed for %s: %s", image_name, exc
            )
        yield ledger.event(
            "tool.finished",
            {
                "tool_name": "insert_image_into_docx",
                "success": success,
                "native_chart_completion": True,
                "result_preview": tool_result_preview(
                    "insert_image_into_docx", result, 1200
                ),
            },
            step_id=step_id,
        )
        if not success:
            continue
        extracted = runtime._extract_file_changes(
            "insert_image_into_docx", tool_args, result
        )
        for change in extracted:
            if not isinstance(change, dict):
                continue
            changes.append(change)
            yield ledger.event("file.changed", change, step_id=step_id)
    return changes


__all__ = [
    "FinancialReportRecoveryResult",
    "financial_chart_recovery_code",
    "financial_chart_recovery_tool_args",
    "insert_pending_generated_docx_images_native",
    "pending_generated_docx_images",
    "recover_financial_report",
]
