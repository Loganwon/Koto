# -*- coding: utf-8 -*-
# Copyright (C) 2024-2026 Koto AI. All rights reserved.
# SPDX-License-Identifier: LicenseRef-Koto-Proprietary
from __future__ import annotations

import io
from typing import Any


def normalize_workbook_export_payload(data: Any) -> tuple[Any, list[Any]]:
    if isinstance(data, dict):
        return data.get("snapshot") or data, data.get("_images", [])
    return data, []


def export_workbook_payload(data: Any) -> bytes:
    workbook_data, images_data = normalize_workbook_export_payload(data)
    return export_xlsx(workbook_data, images_data)


def export_xlsx(workbook_data: Any, images_data: Any | None = None) -> bytes:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 未安装")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _H_ALIGN_REV = {
        0: "general",
        1: "left",
        2: "center",
        3: "right",
        4: "justify",
        5: "centerContinuous",
        6: "distributed",
        7: "distributed",
    }
    _V_ALIGN_REV = {
        1: "top",
        2: "center",
        3: "bottom",
        4: "justify",
        5: "distributed",
    }

    def _resolve_style(cell: dict, top_styles: dict) -> dict | None:
        style = cell.get("s")
        if isinstance(style, (str, int)):
            return top_styles.get(style) or top_styles.get(str(style))
        if isinstance(style, dict):
            return style
        return None

    def _apply_style(oc, style: dict) -> None:
        from openpyxl.styles import Alignment as XlAlignment
        from openpyxl.styles import Font, PatternFill

        font_kw: dict[str, Any] = {}
        if style.get("bl"):
            font_kw["bold"] = True
        if style.get("it"):
            font_kw["italic"] = True
        fs = style.get("fs")
        if fs:
            font_kw["size"] = int(fs)
        cl = style.get("cl", {})
        if isinstance(cl, dict) and cl.get("rgb"):
            color_rgb = cl["rgb"]
            if len(color_rgb) == 7 and color_rgb.startswith("#"):
                font_kw["color"] = color_rgb[1:]
        if font_kw:
            oc.font = Font(**font_kw)

        bg = style.get("bg", {})
        if isinstance(bg, dict) and bg.get("rgb"):
            bg_rgb = bg["rgb"]
            if len(bg_rgb) == 7 and bg_rgb.startswith("#"):
                oc.fill = PatternFill(
                    start_color=bg_rgb[1:],
                    end_color=bg_rgb[1:],
                    fill_type="solid",
                )

        ali_kw: dict[str, Any] = {}
        ht = style.get("ht")
        try:
            ht = int(ht) if ht is not None else None
        except (TypeError, ValueError):
            ht = None
        if ht is not None and ht in _H_ALIGN_REV:
            ali_kw["horizontal"] = _H_ALIGN_REV[ht]
        vt = style.get("vt")
        try:
            vt = int(vt) if vt is not None else None
        except (TypeError, ValueError):
            vt = None
        if vt is not None and vt in _V_ALIGN_REV:
            ali_kw["vertical"] = _V_ALIGN_REV[vt]
        # Univer WrapStrategy.WRAP is 3. Accept 2 as a compatibility value for
        # snapshots created by Koto versions that used the old enum mapping.
        if style.get("tb") in (2, 3) or style.get("wrapText") is True:
            ali_kw["wrap_text"] = True
        if ali_kw:
            oc.alignment = XlAlignment(**ali_kw)

        number_format = style.get("n")
        if isinstance(number_format, dict):
            number_format = number_format.get("pattern") or number_format.get("p")
        if isinstance(number_format, str) and number_format:
            oc.number_format = number_format

    if isinstance(workbook_data, dict) and "sheetOrder" in workbook_data:
        top_styles = workbook_data.get("styles", {})
        sheet_order = workbook_data.get("sheetOrder", [])
        sheets_map = workbook_data.get("sheets", {})
        for sheet_id in sheet_order:
            sheet_data = sheets_map.get(sheet_id, {})
            ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
            cell_data = sheet_data.get("cellData", {})
            for row_key, row_cells in cell_data.items():
                r = int(row_key) + 1
                for col_key, cell in row_cells.items():
                    c = int(col_key) + 1
                    if not cell:
                        continue
                    oc = ws.cell(row=r, column=c)
                    formula = cell.get("f")
                    if isinstance(formula, str) and formula.strip():
                        oc.value = formula if formula.startswith("=") else f"={formula}"
                    elif "v" in cell:
                        oc.value = cell["v"]
                        if cell.get("t") == 3:
                            oc.value = bool(cell["v"])
                    style = _resolve_style(cell, top_styles)
                    if style:
                        _apply_style(oc, style)

            # Univer stores dimensions in CSS pixels; openpyxl uses roughly
            # character widths for columns and points for rows.
            for col_key, col_info in (sheet_data.get("columnData") or {}).items():
                if not isinstance(col_info, dict):
                    continue
                col_index = int(col_key) + 1
                width_px = col_info.get("w")
                if width_px is not None:
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(col_index)
                    ].width = max(0.1, (float(width_px) - 5.0) / 7.0)
                if col_info.get("hd") is True:
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(col_index)
                    ].hidden = True

            for row_key, row_info in (sheet_data.get("rowData") or {}).items():
                if not isinstance(row_info, dict):
                    continue
                row_index = int(row_key) + 1
                height_px = row_info.get("h")
                if height_px is not None:
                    ws.row_dimensions[row_index].height = max(
                        0.1, float(height_px) * 0.75
                    )
                if row_info.get("hd") is True:
                    ws.row_dimensions[row_index].hidden = True

            for merge in sheet_data.get("mergeData", []):
                ws.merge_cells(
                    start_row=merge.get("startRow", 0) + 1,
                    start_column=merge.get("startColumn", 0) + 1,
                    end_row=merge.get("endRow", 0) + 1,
                    end_column=merge.get("endColumn", 0) + 1,
                )

    if not wb.worksheets:
        wb.create_sheet(title="Sheet1")

    if images_data and wb.worksheets:
        import base64

        try:
            from openpyxl.drawing.image import Image as XlImage
        except ImportError:
            XlImage = None
        if XlImage:
            ws_first = wb.worksheets[0]
            for img_data in images_data:
                src = img_data.get("src", "")
                if not src or not src.startswith("data:image"):
                    continue
                try:
                    b64 = src.split(",", 1)[1]
                    raw = base64.b64decode(b64)
                    ximg = XlImage(io.BytesIO(raw))
                    ws_first.add_image(ximg, "A1")
                except Exception:
                    continue

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


__all__ = [
    "export_workbook_payload",
    "export_xlsx",
    "normalize_workbook_export_payload",
]
