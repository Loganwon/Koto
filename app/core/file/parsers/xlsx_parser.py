# -*- coding: utf-8 -*-
# Copyright (C) 2024-2026 Koto AI. All rights reserved.
# SPDX-License-Identifier: LicenseRef-Koto-Proprietary
from __future__ import annotations

import os
import uuid
from typing import Any

_UNIVER_TYPE_STRING = 1
_UNIVER_TYPE_NUMBER = 2
_UNIVER_TYPE_BOOLEAN = 3

_ALIGN_H_MAP = {
    "general": 0,
    "left": 1,
    "center": 2,
    "right": 3,
    "justify": 6,
    "distributed": 7,
}

_ALIGN_V_MAP = {
    "top": 1,
    "middle": 2,
    "center": 2,
    "bottom": 3,
    "justify": 4,
    "distributed": 5,
}


def xlsx_contains_formula_fast(path: str) -> bool:
    import zipfile

    try:
        with zipfile.ZipFile(path, "r") as zf:
            for info in zf.infolist():
                name = info.filename
                if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                    continue
                with zf.open(info) as fh:
                    tail = b""
                    for chunk in iter(lambda: fh.read(65536), b""):
                        data = tail + chunk
                        if b"<f" in data:
                            return True
                        tail = data[-8:]
    except Exception:
        return False
    return False


def openpyxl_cell_to_univer(cell: Any) -> dict[str, Any] | None:
    v = cell.value
    if v is None:
        return None

    cell_data: dict[str, Any] = {}
    if isinstance(v, bool):
        cell_data["v"] = int(v)
        cell_data["t"] = _UNIVER_TYPE_BOOLEAN
    elif isinstance(v, (int, float)):
        cell_data["v"] = v
        cell_data["t"] = _UNIVER_TYPE_NUMBER
    else:
        cell_data["v"] = str(v)
        cell_data["t"] = _UNIVER_TYPE_STRING

    style: dict[str, Any] = {}
    try:
        font = cell.font
        if font:
            if font.bold:
                style["bl"] = 1
            if font.italic:
                style["it"] = 1
            if font.size:
                style["fs"] = int(font.size)
            if (
                font.color
                and font.color.type == "rgb"
                and font.color.rgb
                not in (
                    "00000000",
                    "FF000000",
                )
            ):
                style["cl"] = {"rgb": "#" + font.color.rgb[2:]}
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none") and fill.fgColor:
            if fill.fgColor.type == "rgb" and fill.fgColor.rgb not in (
                "00000000",
                "FFFFFFFF",
            ):
                style["bg"] = {"rgb": "#" + fill.fgColor.rgb[2:]}
        alignment = cell.alignment
        if alignment:
            if alignment.horizontal and alignment.horizontal in _ALIGN_H_MAP:
                style["ht"] = _ALIGN_H_MAP[alignment.horizontal]
            if alignment.vertical and alignment.vertical in _ALIGN_V_MAP:
                style["vt"] = _ALIGN_V_MAP[alignment.vertical]
    except Exception:
        pass

    if style:
        cell_data["s"] = style
    return cell_data


def parse_xlsx(file_path: str, original_name: str | None = None) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 未安装，请执行: pip install openpyxl")

    warnings: list[str] = []
    if xlsx_contains_formula_fast(file_path):
        warnings.append(
            "此表格包含公式（如 =SUM(...)）。Koto 目前以「静态值」模式读取 Excel，"
            "公式已转换为计算结果，保存导出后公式将永久丢失。如需保留公式，请下载原始文件。"
        )

    wb = openpyxl.load_workbook(file_path, data_only=True)

    workbook_id = str(uuid.uuid4())
    if original_name:
        workbook_name = os.path.splitext(os.path.basename(original_name))[0]
    else:
        workbook_name = os.path.splitext(os.path.basename(file_path))[0]
    sheet_order: list[str] = []
    sheets: dict[str, Any] = {}
    style_hash_to_id: dict[str, str] = {}
    styles_registry: dict[str, Any] = {}

    def _get_style_id(style_obj: dict[str, Any]) -> str:
        import json as _json

        style_hash = _json.dumps(style_obj, sort_keys=True, ensure_ascii=False)
        if style_hash not in style_hash_to_id:
            style_id = str(len(style_hash_to_id))
            style_hash_to_id[style_hash] = style_id
            styles_registry[style_id] = style_obj
        return style_hash_to_id[style_hash]

    for idx, ws in enumerate(wb.worksheets):
        sheet_id = f"sheet{idx + 1}"
        sheet_order.append(sheet_id)

        cell_data: dict[int, dict[int, Any]] = {}
        for row in ws.iter_rows():
            for cell in row:
                converted = openpyxl_cell_to_univer(cell)
                if converted is None:
                    continue
                if "s" in converted and isinstance(converted["s"], dict):
                    converted["s"] = _get_style_id(converted["s"])
                row_idx = cell.row - 1
                col_idx = cell.column - 1
                cell_data.setdefault(row_idx, {})[col_idx] = converted

        merge_data: list[dict[str, int]] = []
        for merge_range in ws.merged_cells.ranges:
            merge_data.append(
                {
                    "startRow": merge_range.min_row - 1,
                    "startColumn": merge_range.min_col - 1,
                    "endRow": merge_range.max_row - 1,
                    "endColumn": merge_range.max_col - 1,
                }
            )

        column_data: dict[int, dict] = {}
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width and col_dim.width > 0:
                col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                column_data[col_idx] = {"w": max(4, round(col_dim.width * 7 + 5))}

        row_data: dict[int, dict] = {}
        for row_num, row_dim in ws.row_dimensions.items():
            if row_dim.height and row_dim.height > 0:
                row_data[row_num - 1] = {"h": max(4, round(row_dim.height * 96 / 72))}

        sheets[sheet_id] = {
            "id": sheet_id,
            "name": ws.title,
            "rowCount": max(ws.max_row or 30, 30),
            "columnCount": max(ws.max_column or 10, 10),
            "cellData": cell_data,
            "mergeData": merge_data,
            "columnData": column_data,
            "rowData": row_data,
        }

    wb.close()

    return {
        "id": workbook_id,
        "name": workbook_name,
        "appVersion": "0.5.0",
        "locale": "zh-CN",
        "sheetOrder": sheet_order,
        "sheets": sheets,
        "styles": styles_registry,
        "resources": [],
        "_warnings": warnings,
    }


__all__ = ["openpyxl_cell_to_univer", "parse_xlsx", "xlsx_contains_formula_fast"]
